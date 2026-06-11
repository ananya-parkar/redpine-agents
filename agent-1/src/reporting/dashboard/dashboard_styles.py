# agent-1/src/reporting/dashboard/dashboard_styles.py
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.reporting.dashboard.dashboard_constants import *

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=TEXT_DARK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Arial")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _apply(ws, rng, **kw):
    for row in ws[rng]:
        for c in row:
            for k, v in kw.items():
                setattr(c, k, v)

def _section_title(ws, rng, title, bg=GREEN_PRIMARY):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = title
    c.font  = _font(bold=True, size=10, color=WHITE)
    c.fill  = _fill(bg)
    c.alignment = _align("left", "center")
    _apply(ws, rng, fill=_fill(bg))
