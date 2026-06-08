# agent-1/src/dashboard/dashboard_sheets.py
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from src.dashboard.dashboard_constants import *
from src.dashboard.dashboard_styles import _fill, _font, _align, _border, _apply, _section_title


def _sheet_header(ws, headers, widths=None):
    ws.row_dimensions[1].height = 22
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i+1, value=h)
        c.font      = Font(bold=True, size=10, color=WHITE, name="Arial")
        c.fill      = _fill(GREEN_PRIMARY)
        c.alignment = _align("center")
        c.border    = _border("FFFFFF")
        if widths and i < len(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = widths[i]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def _write_rows(ws, df, columns, start_row=2):
    for ri, (_, row) in enumerate(df.iterrows(), start=start_row):
        ws.row_dimensions[ri].height = 18
        for ci, col in enumerate(columns, start=1):
            val = row.get(col, "")
            if isinstance(val, float) and pd.isna(val): val = ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(size=9, color=TEXT_DARK, name="Arial")
            c.border = _border()
            c.alignment = _align("left" if ci == 1 else "center")
            if ri % 2 == 0: c.fill = _fill("F8FAFC")

def populate_all_leads(ws, df):
    col_defs = [
        ("hotel_name","Hotel Name",22), ("city","City",14), ("state","State",7),
        ("final_lead_score","Lead Score",11), ("opportunity_score","Opp. Score",11),
        ("llm_star_rating","AI Rating",10), ("distress_probability","Distress Prob.",13),
        ("owner_name","Owner Name",22), ("ownership_length_years","Own. Years",10),
        ("property_age","Prop. Age",10), ("room_count","Rooms",8),
        ("current_brand","Brand",14), ("franchise_affiliated","Franchise",10),
        ("cmbs_watchlist","CMBS Watch",11), ("cmbs_delinquent","CMBS Delq.",11),
        ("cmbs_special_servicing","Special Svc.",11),
        ("signals_fired","Signals Fired",35),
        ("lead_status","Lead Status ▼",14), ("notes","Notes",30),
        ("created_at","Date First Surfaced",18),
    ]
    for i, (_, lbl, w) in enumerate(col_defs):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.row_dimensions[1].height = 22
    for i, (_, lbl, _) in enumerate(col_defs):
        c = ws.cell(row=1, column=i+1, value=lbl)
        c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
        c.fill = _fill(GREEN_PRIMARY); c.alignment = _align("center")
        c.border = _border("FFFFFF")

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[ri].height = 18
        for ci, (key, _, _) in enumerate(col_defs, start=1):
            val = row.get(key, "")
            if isinstance(val, float) and pd.isna(val): val = ""
            if key == "distress_probability" and val != "": val = round(float(val), 2)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=9, color=TEXT_DARK, name="Arial")
            c.alignment = _align("left" if ci in [1,2,8,17,19] else "center")
            c.border = _border()
            if ri % 2 == 0: c.fill = _fill("F8FAFC")
            if key == "lead_status":
                sc_bg, sc_fg = STATUS_COLORS.get(str(val), ("E2E8F0", TEXT_DARK))
                c.font = Font(bold=True, size=9, color=sc_fg, name="Arial")
                c.fill = _fill(sc_bg)
            if key == "final_lead_score" and val != "":
                sv = float(val)
                if sv >= 80:   c.fill=_fill("DCFCE7"); c.font=Font(bold=True,size=9,color="166534",name="Arial")
                elif sv >= 60: c.fill=_fill("FEF9C3"); c.font=Font(bold=True,size=9,color="854D0E",name="Arial")
                elif sv < 50:  c.fill=_fill("FEE2E2"); c.font=Font(bold=True,size=9,color="991B1B",name="Arial")

    # Status dropdown
    status_col = [k for k,_,_ in col_defs].index("lead_status") + 1
    sc_letter  = get_column_letter(status_col)
    dv = DataValidation(type="list",
                        formula1='"New,Pursuing,Monitoring,Passed,Underwriting"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{sc_letter}2:{sc_letter}{len(df)+1}")

    # Score gradient
    score_letter = get_column_letter([k for k,_,_ in col_defs].index("final_lead_score")+1)
    ws.conditional_formatting.add(
        f"{score_letter}2:{score_letter}{len(df)+1}",
        ColorScaleRule(start_type="min",start_color="FEE2E2",
                       mid_type="percentile",mid_value=50,mid_color="FEF9C3",
                       end_type="max",end_color="DCFCE7")
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def populate_high_opportunity(ws, df):
    high = df[df["final_lead_score"] >= 70].sort_values("final_lead_score", ascending=False)
    cols = ["hotel_name","city","state","final_lead_score","opportunity_score",
            "owner_name","ownership_length_years","signals_fired","lead_status"]
    _sheet_header(ws,["Hotel Name","City","State","Lead Score","Opp. Score",
                       "Owner","Own. Years","Signals","Status"],[24,14,7,11,11,22,11,35,13])
    _write_rows(ws, high, cols)

def populate_ownership_analysis(ws, df):
    od   = df.sort_values("ownership_length_years", ascending=False, na_position="last")
    cols = ["hotel_name","owner_name","ownership_since","ownership_length_years",
            "property_age","final_lead_score","lead_status"]
    _sheet_header(ws,["Hotel Name","Owner","Owned Since","Years Owned","Prop. Age",
                       "Lead Score","Status"],[24,22,13,11,10,11,13])
    _write_rows(ws, od, cols)

def populate_cmbs_sheet(ws, df):
    cmbs = df[df["cmbs_watchlist"].fillna(False)|df["cmbs_delinquent"].fillna(False)|
              df["cmbs_special_servicing"].fillna(False)]
    cols = ["hotel_name","city","state","final_lead_score","cmbs_watchlist",
            "cmbs_delinquent","cmbs_special_servicing","lead_status"]
    _sheet_header(ws,["Hotel Name","City","State","Lead Score","Watchlist",
                       "Delinquent","Special Svc.","Status"],[24,14,7,11,12,12,14,13])
    _write_rows(ws, cmbs, cols)

def populate_distress_sheet(ws, df):
    dd   = df.sort_values("distress_probability", ascending=False)
    cols = ["hotel_name","signals_fired","distress_probability",
            "seller_fatigue_probability","final_lead_score","lead_status"]
    _sheet_header(ws,["Hotel Name","Signals","Distress Prob.","Seller Fatigue",
                       "Lead Score","Status"],[24,35,14,18,11,13])
    _write_rows(ws, dd, cols)

def populate_lead_tracker(ws, df):
    cols   = ["hotel_name","city","state","final_lead_score","lead_status","notes","created_at"]
    widths = [24,14,7,11,13,35,18]
    _sheet_header(ws,["Hotel Name","City","State","Lead Score","Status ▼",
                       "Notes","Date Surfaced"], widths)
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[ri].height = 18
        for ci, key in enumerate(cols, start=1):
            val = row.get(key,"")
            if isinstance(val, float) and pd.isna(val): val = ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(size=9, color=TEXT_DARK, name="Arial")
            c.border = _border()
            c.alignment = _align("left" if ci in [1,6] else "center")
            if ri % 2 == 0: c.fill = _fill("F8FAFC")
            if key == "lead_status":
                sc_bg, sc_fg = STATUS_COLORS.get(str(val),("E2E8F0",TEXT_DARK))
                c.font = Font(bold=True,size=9,color=sc_fg,name="Arial")
                c.fill = _fill(sc_bg)
    dv = DataValidation(type="list",
                        formula1='"New,Pursuing,Monitoring,Passed,Underwriting"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"E2:E{len(df)+1}")
    note_r = len(df) + 3
    ws.merge_cells(f"A{note_r}:G{note_r}")
    nc = ws.cell(row=note_r, column=1,
                 value="ℹ️  Update Status or Notes above — agent reads this sheet to deduplicate and adjust scoring.")
    nc.font      = Font(size=9, color="1E40AF", italic=True, name="Arial")
    nc.fill      = _fill("EFF6FF")
    nc.alignment = _align("left")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
