import json
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import dashboard writer (your existing dashboard code)
from dashboard_writer import write_dashboard

# Import DB loader for dashboard historical data
from db_writer import get_all_signals_for_excel, get_all_leads_for_excel, get_pending_tuning_triggers
from tuning_prompt import _match_rule

# ---------------------------------------------------
# EXCEL BUILDER — 5 sheets
# 1. Dashboard   ← loads from PostgreSQL (historical)
# 2. Venue Database
# 3. All Leads
# 4. Act Now     (subset of All Leads)
# 5. Stakeholders
# ---------------------------------------------------

TIER_BG = {1:"C8E6C9",2:"BBDEFB",3:"FFE0B2",4:"FFCDD2"}
TIER_FG = {1:"1B5E20",2:"0D47A1",3:"BF360C",4:"B71C1C"}
ENG_BG  = {
    "engage_now": ("C8E6C9","1B5E20"),
    "monitor":    ("FFF9C4","E65100"),
    "too_late":   ("FFCDD2","B71C1C"),
}

def fill(h):  return PatternFill("solid",start_color=h,end_color=h)
def bdr():
    s=Side(style="thin",color="DDDDDD")
    return Border(left=s,right=s,top=s,bottom=s)
def al(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def fnt(color="212121",bold=False,sz=9):
    return Font(color=color,bold=bold,size=sz,name="Calibri")
def stripe(i): return fill("F5F5F5" if i%2==0 else "FFFFFF")

def header_row(ws, headers, widths, bg):
    for col,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=1,column=col,value=h)
        c.font=Font(bold=True,color="FFFFFF",size=10,name="Calibri")
        c.fill=fill(bg); c.alignment=al(); c.border=bdr()
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[1].height=28
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}1"

    # Feedback dropdown (col 17) — Pursuing / Archive / Bad Data / Passed / Watch
    # NOTE: "Watch" was missing here even though feedback_reader.py
    # already handles a "watch" feedback value — without it in the
    # dropdown list, nobody could actually select it (short of typing
    # it manually and bypassing the dropdown). Added below so the two
    # stay in sync.
    feedback_col = None
    for idx, h in enumerate(headers, 1):
        if h == "Feedback":
            feedback_col = idx
            break
    if feedback_col:
        dv = DataValidation(
            type="list",
            formula1='"Pursuing,Archive,Bad Data,Passed,Watch,"',
            allow_blank=True,
            showDropDown=False,   # False = arrow IS shown (Excel inverted logic)
        )
        dv.sqref = f"{get_column_letter(feedback_col)}2:{get_column_letter(feedback_col)}500"
        ws.add_data_validation(dv)


# ── SHEET: VENUE DATABASE ────────────────────────────────────────
def write_venue_db(ws, venues):
    ws.title = "Venue Database"
    headers = ["Venue Name","League","Team","City","State","Capacity",
               "Year Built / Planned","Status","Owner","Operator",
               "Facilities Contact"]
    widths  = [34,16,28,18,5,10,16,28,30,30,25]
    header_row(ws, headers, widths, "004D40")
    for i,v in enumerate(venues,2):
        status = v.get("status","existing")
        yr = v.get("year_built","") or v.get("planned_year","")
        row = [v.get("venue_name",""), v.get("league",""), v.get("team",""),
               v.get("city",""), v.get("state",""), v.get("capacity","") or "",
               yr, status, v.get("owner",""), v.get("operator",""),
               v.get("facilities_contact","")]
        for col,val in enumerate(row,1):
            c=ws.cell(row=i,column=col,value=val)
            c.border=bdr(); c.font=fnt()
            c.alignment=al("left" if col in [1,3,4] else "center")
            if col==8 and "planned" in str(status):
                c.fill=fill("C8E6C9")
                c.font=Font(bold=True,color="1B5E20",size=9,name="Calibri")
            else: c.fill=stripe(i)
        ws.row_dimensions[i].height=18


# ── SHARED LEADS WRITER ──────────────────────────────────────────
LEAD_HEADERS = [
    "Rank","Venue","League","Team","City","State","Capacity",
    "Signal Tier","Score","Likelihood %","Project Type",
    "What's Happening","Why Priority","Evidence","Stakeholders",
    "Source (Link)","Engagement","Feedback","Notes"
]
LEAD_WIDTHS = [6,28,8,22,16,5,10,26,8,12,18,46,36,40,35,40,14,14,25]
LEAD_LEFT   = {2,4,5,12,13,14,15,16}

def _lead_row_vals(r: dict, rank_field: str = "rank") -> list:
    """Extract display values from a lead record."""
    t   = r.get("signal_tier") or 0
    eng = r.get("engagement") or r.get("engagement_action","monitor")
    scr = r.get("score") or r.get("final_score","")
    lkl = r.get("likelihood","")

    # Stakeholders text
    try:
        raw = r.get("stakeholders_raw") or r.get("stakeholder_names","") or "[]"
        if raw.startswith("["):
            stakes = json.loads(raw)
            stakes_str = "\n".join(
                f"{s.get('name','')} [{s.get('type','')}]"
                for s in stakes if s.get("name")
            )
        else:
            stakes_str = raw  # already a semicolon-separated string from DB
    except Exception:
        stakes_str = ""

    return [
        r.get(rank_field, ""),
        r.get("venue_name",""),
        r.get("league",""),
        r.get("team",""),
        r.get("city",""),
        r.get("state",""),
        r.get("capacity",""),
        r.get("tier_label","") or (f"Tier {t}" if t else ""),
        scr,
        f"{lkl}%" if lkl else "",
        r.get("project_type",""),
        r.get("whats_happening",""),
        r.get("why_priority",""),
        r.get("evidence",""),
        stakes_str,
        r.get("source",""),
        eng,
        r.get("feedback",""),
        r.get("notes",""),
    ]

def write_leads_sheet(ws, results, header_bg, rank_field="rank"):
    header_row(ws, LEAD_HEADERS, LEAD_WIDTHS, header_bg)
    for i,r in enumerate(results,2):
        t   = r.get("signal_tier") or 0
        eng = r.get("engagement") or r.get("engagement_action","monitor")
        ebg,efg = ENG_BG.get(eng,("FFFFFF","000000"))
        sbg = "F9FAFB" if i%2==0 else "FFFFFF"
        row = _lead_row_vals(r, rank_field)
        for col,val in enumerate(row,1):
            c=ws.cell(row=i,column=col,value=val)
            c.border=bdr()
            c.alignment=al("left" if col in LEAD_LEFT else "center",
                           wrap=(col in {12,13,14,15}))
            if col==8 and t and t in TIER_BG:
                c.fill=fill(TIER_BG[t])
                c.font=Font(bold=True,color=TIER_FG[t],size=9,name="Calibri")
            elif col==17:
                c.fill=fill(ebg)
                c.font=Font(bold=True,color=efg,size=9,name="Calibri")
            else:
                c.fill=fill(sbg); c.font=fnt()
        ws.row_dimensions[i].height=75


# ── SHEET: STAKEHOLDERS ──────────────────────────────────────────
def write_stakeholders(ws, stakeholder_rows):
    ws.title="Stakeholders"
    headers=["Venue","League","Team","Signal Tier","Engagement",
             "Name","Title","Organization","Type","Website",
             "Contact Email","Notes"]
    widths=[28,8,22,24,14,28,22,28,14,25,28,25]
    header_row(ws,headers,widths,"4A148C")
    for i,s in enumerate(stakeholder_rows,2):
        t   = s.get("signal_tier") or 0
        eng = s.get("engagement") or s.get("engagement_action","")
        ebg,efg=ENG_BG.get(eng,("FFFFFF","000000"))
        sbg="FAF5FF" if i%2==0 else "FFFFFF"
        row=[s.get("venue_name",""),s.get("league",""),s.get("team",""),
             f"Tier {t}" if t else "",eng,
             s.get("stakeholder_name",""),s.get("title",""),
             s.get("organization",""),s.get("type",""),
             s.get("website",""),s.get("contact_email",""),s.get("notes","")]
        LEFT={1,3,6,7,8}
        for col,val in enumerate(row,1):
            c=ws.cell(row=i,column=col,value=val); c.border=bdr()
            c.alignment=al("left" if col in LEFT else "center")
            if col==4 and t and t in TIER_BG:
                c.fill=fill(TIER_BG[t]); c.font=Font(bold=True,color=TIER_FG[t],size=9,name="Calibri")
            elif col==5:
                c.fill=fill(ebg); c.font=Font(bold=True,color=efg,size=9,name="Calibri")
            else:
                c.fill=fill(sbg); c.font=fnt()
        ws.row_dimensions[i].height=20


# ── SHEET: TUNING REVIEW ─────────────────────────────────────────
SCOPE_LABELS = {
    "reasoning":   "Tier/Relevance Logic",
    "stakeholder": "Stakeholder Extraction",
    "code_only":   "⚠️ Code Fix Needed (not an LLM prompt)",
}

def write_tuning_review(ws, active_triggers):
    """
    Read-only audit trail of currently active tuning patterns — which
    Bad Data notes were flagged 3+ times, what instruction got applied,
    and to which prompt. Per the client's requirement, these activate
    automatically (3+ same note = trigger) with no separate approval
    step, so this sheet is for transparency only — nothing to fill in.
    """
    ws.title = "Tuning Review"
    headers = ["ID","Root Cause","Occurrences","Example Venues",
               "Target","Applied Instruction","Status"]
    widths  = [6,22,12,32,24,55,12]
    header_row(ws, headers, widths, "5D4037")

    for i, t in enumerate(active_triggers, 2):
        root = t.get("root_cause") or "unspecified"
        rule = _match_rule(root)
        target = SCOPE_LABELS.get(rule["scope"], rule["scope"])
        instruction = rule.get("instruction") or "(no LLM instruction — needs a manual code/data fix instead)"
        row = [
            t.get("id",""), root, t.get("occurrences",""),
            (t.get("affected_venues") or "")[:150],
            target, instruction,
            "Active" if rule["scope"] != "code_only" else "Logged (code fix needed)",
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = bdr(); c.font = fnt()
            c.alignment = al("left" if col in (2,4,6) else "center", wrap=(col in (4,6)))
            if col == 5 and rule["scope"] == "code_only":
                c.fill = fill("FFF3E0")  # amber tint — flags "not LLM-fixable"
                c.font = Font(bold=True, color="E65100", size=9, name="Calibri")
            elif col == 7:
                if rule["scope"] == "code_only":
                    c.fill = fill("FFF3E0"); c.font = Font(bold=True, color="E65100", size=9, name="Calibri")
                else:
                    c.fill = fill("C8E6C9"); c.font = Font(bold=True, color="1B5E20", size=9, name="Calibri")
            else:
                c.fill = fill("F5F5F5" if i % 2 == 0 else "FFFFFF")
        ws.row_dimensions[i].height = 42

    if not active_triggers:
        ws.cell(row=2, column=1, value="No recurring patterns flagged yet.").font = fnt()


# ── MAIN SAVE ────────────────────────────────────────────────────
def save_excel(venues, all_leads, act_now, stakeholder_rows,
               output_path, run_date=None):
    print("\n[STEP 5] Writing Excel...", flush=True)
    run_date = run_date or date.today()
    wb = Workbook()

    # ── Sheet 1: Dashboard (reads from PostgreSQL) ───────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    try:
        # Load from DB — last 90 days of signals + today's leads
        db_signals = get_all_signals_for_excel(days=90)
        db_leads   = get_all_leads_for_excel()

        # FIX: signals.signal_tier is always NULL in the DB — it only
        # gets computed by the LLM and saved to the LEADS table, never
        # written back to the signals table. dashboard_writer's tier
        # KPI/ring-chart reads from the signals list, so without this
        # patch "Tier 1 Signals" etc. always show 0. Patch each signal's
        # tier from the matching lead (same venue) before charting —
        # dashboard_writer.py itself stays untouched.
        tier_by_venue = {l["venue_name"]: l.get("signal_tier")
                         for l in db_leads if l.get("venue_name")}
        for s in db_signals:
            if s.get("signal_tier") is None:
                matched_tier = tier_by_venue.get(s.get("venue_name"))
                if matched_tier is not None:
                    s["signal_tier"] = matched_tier

        # DB leads already have correct field names (engagement_action, final_score etc.)
        # write_dashboard uses: engagement_action, final_score, signal_tier, signal_type
        write_dashboard(ws_dash, db_signals, db_leads)
        print("  Dashboard loaded from PostgreSQL ✅", flush=True)
    except Exception as e:
        print(f"  [WARN] Dashboard from DB failed ({e}), using current run data", flush=True)
        # Fallback: use current run data with field name mapping
        mapped_leads = []
        for r in all_leads:
            ml = dict(r)
            ml["engagement_action"] = r.get("engagement","monitor")
            ml["final_score"]       = r.get("score","")
            mapped_leads.append(ml)
        mapped_signals = []
        for s in (all_leads or []):
            ms = dict(s)
            ms["signal_type"] = "news"
            mapped_signals.append(ms)
        try:
            write_dashboard(ws_dash, mapped_signals, mapped_leads)
        except Exception as e2:
            # Last-resort guard: never let dashboard rendering crash the
            # whole Excel export — write a simple placeholder instead.
            print(f"  [WARN] Dashboard fallback also failed ({e2}) — "
                  f"writing placeholder sheet", flush=True)
            try:
                # Unmerge any ranges left over from the partial write,
                # otherwise clearing cell values below will also fail
                # with "MergedCell is read-only".
                for merged_range in list(ws_dash.merged_cells.ranges):
                    ws_dash.unmerge_cells(str(merged_range))
                for row in ws_dash.iter_rows():
                    for cell in row:
                        cell.value = None
            except Exception:
                pass  # if even cleanup fails, just leave whatever's there
            ws_dash["B2"] = "Dashboard unavailable for this run (insufficient data)."
            ws_dash["B3"] = "All Leads / Act Now / Venue Database sheets are unaffected."

    # ── Sheet 2: Venue Database ──────────────────────────────────
    ws_venues = wb.create_sheet()
    write_venue_db(ws_venues, venues)

    # ── Sheet 3: All Leads ───────────────────────────────────────
    ws_all = wb.create_sheet()
    ws_all.title = "All Leads"
    write_leads_sheet(ws_all, all_leads, "1A237E", rank_field="rank")

    # ── Sheet 4: Act Now ────────────────────────────────────────
    ws_act = wb.create_sheet()
    ws_act.title = "Act Now"
    write_leads_sheet(ws_act, act_now, "B71C1C", rank_field="act_now_rank")

    # ── Sheet 5: Stakeholders ────────────────────────────────────
    ws_stakes = wb.create_sheet()
    write_stakeholders(ws_stakes, stakeholder_rows)

    # ── Sheet 6: Tuning Review ───────────────────────────────────
    # Pending feedback-driven patterns awaiting Approve/Reject — see
    # write_tuning_review() above. Fetched fresh each run so the sheet
    # always reflects current DB state (cleared patterns disappear,
    # new ones appear).
    ws_tuning = wb.create_sheet()
    pending_triggers = get_pending_tuning_triggers()
    write_tuning_review(ws_tuning, pending_triggers)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Saved → {output_path}", flush=True)
    print(f"  Venue Database  : {len(venues)}", flush=True)
    print(f"  All Leads       : {len(all_leads)}", flush=True)
    print(f"  Act Now         : {len(act_now)}", flush=True)
    print(f"  Stakeholders    : {len(stakeholder_rows)}", flush=True)