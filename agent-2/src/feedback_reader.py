# ---------------------------------------------------
# FEEDBACK READER  v2
# Reads Matthew's feedback from All Leads Excel sheet
# → writes to DB → detects bad-data patterns → logs tuning triggers
#
# Feedback tags and what they mean:
#   Pursuing   → score +15, retain PII, force to Act Now
#   Passed     → valid lead, Matthew chose not to pursue
#   Bad Data   → false positive; Notes = root cause
#                 3+ same root cause → tuning trigger logged
#   Watch      → keep monitoring
#   Archive    → exclude from future pipeline runs
#
# UPDATED: reads from the latest PREVIOUS dated run file in runs/
# (see run_paths.py) instead of a single fixed OUTPUT_FILE — since
# each run now writes its own dated Excel, the file someone actually
# edited is whichever one was most recently created before today's.
# ---------------------------------------------------

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook
from db_writer  import get_conn, get_bad_data_patterns, log_tuning_trigger
from run_paths  import get_latest_previous_run_file
from datetime   import datetime, timezone

FEEDBACK_COL = 17   # col Q in All Leads
VENUE_COL    = 2    # col B
NOTES_COL    = 18   # col R  ← root cause goes here for Bad Data

VALID_FEEDBACK = {
    "pursuing", "passed", "bad data",
    "watch", "archive"
}

# Maps root cause keyword → what to fix in the pipeline
TUNING_RECOMMENDATIONS = {
    "wrong owner":       "Review stakeholder_enrichment.py — owner matching logic may be too loose.",
    "wrong venue":       "Review venue_fetcher.py — venue name matching may have false matches.",
    "no construction":   "Review heuristic_scoring.py — keyword list may be triggering on non-construction articles.",
    "old news":          "Review signal_collector.py — increase recency filter, reduce lookback window.",
    "not our market":    "Review venue_fetcher.py — league/venue filter may be too broad.",
    "duplicate":         "Review deduplicate() in main.py — dedup logic missing some cases.",
    "fake signal":       "Review reasoning_agent.py — LLM prompt needs stronger false-positive filter.",
}

def _get_recommendation(root_cause: str) -> str:
    """Match root cause to a known fix recommendation."""
    rc_lower = root_cause.lower()
    for keyword, rec in TUNING_RECOMMENDATIONS.items():
        if keyword in rc_lower:
            return rec
    return (f"Review pipeline for root cause: '{root_cause}'. "
            f"Check heuristic_scoring.py and reasoning_agent.py prompts.")


def read_feedback_from_excel() -> list[dict]:
    target_file = get_latest_previous_run_file()
    if target_file is None:
        print("[FEEDBACK] No previous run file found yet (first run?) — skipping")
        return []
    if not target_file.exists():
        print(f"[FEEDBACK] Excel not found: {target_file}")
        return []

    print(f"[FEEDBACK] Reading from: {target_file.name}")
    wb = load_workbook(target_file, read_only=True, data_only=True)
    if "All Leads" not in wb.sheetnames:
        print("[FEEDBACK] 'All Leads' sheet not found")
        wb.close()
        return []

    ws = wb["All Leads"]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        venue    = row[VENUE_COL - 1]
        feedback = row[FEEDBACK_COL - 1]
        note     = row[NOTES_COL - 1]

        if not venue or not feedback:
            continue
        fb_clean = str(feedback).strip().lower()
        if fb_clean not in VALID_FEEDBACK:
            continue

        entries.append({
            "venue_name":     str(venue).strip(),
            "feedback":       str(feedback).strip(),
            "feedback_lower": fb_clean,
            "note":           str(note).strip() if note else "",
        })

    wb.close()
    print(f"[FEEDBACK] Found {len(entries)} feedback entries in Excel")
    return entries


def apply_feedback_to_db(entries: list[dict]) -> dict:
    if not entries:
        return {"updated": 0, "bad_data": 0, "pursuing": 0, "passed": 0}

    stats = {"updated": 0, "bad_data": 0, "pursuing": 0, "passed": 0, "errors": 0}
    now   = datetime.now(timezone.utc)

    try:
        conn = get_conn()
        cur  = conn.cursor()
    except Exception as e:
        print(f"[FEEDBACK] DB connection failed: {e}")
        return stats

    for entry in entries:
        venue = entry["venue_name"]
        fb    = entry["feedback"]
        fb_l  = entry["feedback_lower"]
        note  = entry["note"]

        try:
            cur.execute("""
                UPDATE leads
                SET feedback      = %s,
                    feedback_note = %s,
                    feedback_at   = %s,
                    retain_pii    = %s
                WHERE LOWER(venue_name) = LOWER(%s)
            """, (
                fb, note, now,
                fb_l == "pursuing",
                venue
            ))

            if cur.rowcount == 0:
                continue

            stats["updated"] += 1

            if fb_l == "pursuing":
                stats["pursuing"] += 1
                print(f"  ✅ PURSUING    {venue[:45]}", flush=True)

            elif fb_l == "passed":
                stats["passed"] += 1
                print(f"  ⏭  PASSED      {venue[:45]}", flush=True)

            elif fb_l == "bad data":
                stats["bad_data"] += 1
                root_cause = note or "unspecified"
                print(f"  ❌ BAD DATA    {venue[:38]}  [{root_cause}]", flush=True)
                # Archive bad data leads — never show again
                cur.execute("""
                    UPDATE leads SET current_engagement = 'archived'
                    WHERE LOWER(venue_name) = LOWER(%s)
                """, (venue,))

            elif fb_l == "archive":
                cur.execute("""
                    UPDATE leads SET current_engagement = 'archived'
                    WHERE LOWER(venue_name) = LOWER(%s)
                """, (venue,))
                print(f"  🗑  ARCHIVED    {venue[:45]}", flush=True)

            elif fb_l == "watch":
                print(f"  👀 WATCH       {venue[:45]}", flush=True)

            conn.commit()

        except Exception as e:
            conn.rollback()
            print(f"  [ERROR] {venue}: {e}")
            stats["errors"] += 1

    cur.close()
    conn.close()
    return stats


def check_and_log_tuning_triggers():
    """
    Check if 3+ leads share the same Bad Data root cause.
    If yes: log a tuning trigger with a specific fix recommendation.
    This acts as an alert to the developer to fix that part of the pipeline.
    """
    patterns = get_bad_data_patterns()
    if not patterns:
        return

    print(f"\n  ⚠️  TUNING TRIGGERS DETECTED:")
    for p in patterns:
        root_cause = p["root_cause"] or "unspecified"
        count      = p["occurrences"]
        venues     = p["affected_venues"]
        rec        = _get_recommendation(root_cause)

        print(f"\n  Root cause : '{root_cause}'")
        print(f"  Count      : {count} leads")
        print(f"  Venues     : {venues[:80]}")
        print(f"  ACTION     : {rec}")

        log_tuning_trigger(root_cause, count, venues, rec)

    # Write tuning report to a local file for developer reference
    report_path = Path(__file__).parent / "tuning_triggers.txt"
    with open(report_path, "a") as f:
        f.write(f"\n{'='*60}\n")
        f.write(f"TUNING TRIGGERS — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"{'='*60}\n")
        for p in patterns:
            root_cause = p["root_cause"] or "unspecified"
            rec = _get_recommendation(root_cause)
            f.write(f"\nRoot cause  : {root_cause}\n")
            f.write(f"Occurrences : {p['occurrences']}\n")
            f.write(f"Venues      : {p['affected_venues']}\n")
            f.write(f"Fix         : {rec}\n")
    print(f"\n  Tuning report saved → {report_path}")


def run():
    print("\n" + "=" * 55)
    print("  FEEDBACK READER")
    print("=" * 55)

    entries = read_feedback_from_excel()
    if not entries:
        print("  No feedback to process.")
        return

    stats = apply_feedback_to_db(entries)

    print(f"\n  Updated    : {stats['updated']}")
    print(f"  Pursuing   : {stats['pursuing']}")
    print(f"  Passed     : {stats['passed']}")
    print(f"  Bad Data   : {stats['bad_data']}")

    # Check for patterns in bad data → trigger tuning alerts
    if stats["bad_data"] > 0:
        check_and_log_tuning_triggers()

    print("=" * 55)


if __name__ == "__main__":
    run()