# agent-1/src/reporting/dashboard/dashboard_charts.py
import re, zipfile
import pandas as pd
from openpyxl.chart import DoughnutChart, BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from src.reporting.dashboard.dashboard_constants import *
from src.reporting.dashboard.dashboard_metrics import build_signal_summary
from src.reporting.dashboard.dashboard_components import _font, _fill, _align, _apply
from openpyxl.styles import Font, Alignment, PatternFill

# Helpers
# def _color_series(series, colors):
#     for i, hex_color in enumerate(colors):
#         pt = DataPoint(idx=i, invertIfNegative=False)
#         pt.graphicalProperties.solidFill = hex_color
#         series.dPt.append(pt)

def _dLbls(show_pct=False, show_val=False):
    d = DataLabelList()
    d.showPercent     = show_pct
    d.showVal         = show_val
    d.showCatName     = False
    d.showSerName     = False
    d.showLeaderLines = False
    return d


def _write_chart_data(data_ws, df):
    # ── Opportunity distribution (col 1-2, rows 80-85) ──
    dist = [
        ("80-100 Deep",    len(df[df["final_lead_score"] >= 80])),
        ("60-79 High",     len(df[(df["final_lead_score"] >= 60) & (df["final_lead_score"] < 80)])),
        ("40-59 Moderate", len(df[(df["final_lead_score"] >= 40) & (df["final_lead_score"] < 60)])),
        ("20-39 Low",      len(df[(df["final_lead_score"] >= 20) & (df["final_lead_score"] < 40)])),
        ("0-19 Very Low",  len(df[df["final_lead_score"] < 20])),
    ]
    data_ws.cell(100, 1, "Category"); data_ws.cell(100, 2, "Count")
    for i, (label, cnt) in enumerate(dist):
        data_ws.cell(101 + i, 1, label)
        data_ws.cell(101 + i, 2, cnt)

    # ── Distress signals (col 4-5, rows 80-85) ──
    total_hotels = max(len(df), 1)

    sig = {
        k: round((v / total_hotels) * 100)
        for k, v in build_signal_summary(df).items()
    }
    
    sig = dict(
        sorted(
            sig.items(),
            key=lambda x: x[1],
        )
    )
    data_ws.cell(100, 4, "Signal"); data_ws.cell(100, 5, "Count")
    for i, (label, cnt) in enumerate(sig.items()):
        data_ws.cell(101 + i, 4, label)
        data_ws.cell(101 + i, 5, cnt)

    # ── Markets (col 7-8, rows 80-85) ──
    if "city" in df.columns:
    
        market_df = df[
            df["city"].notna()
            & (df["city"].astype(str).str.strip() != "")
            & (~df["city"].astype(str).str.contains(r"\d", na=False))
        ]
    
        mkt = (
            market_df.groupby("city")["final_lead_score"]
            .mean()
            .sort_values(ascending=False)
            .head(5)
        )
        mkt = mkt.iloc[::-1]
    
    else:
        mkt = pd.Series(dtype=float)
    data_ws.cell(100, 7, "Market"); data_ws.cell(100, 8, "Avg Score")
    for i, (market, score) in enumerate(mkt.items()):
        data_ws.cell(101 + i, 7, str(market)[:20])
        data_ws.cell(101 + i, 8, round(float(score), 1))
    
    # - Seller fatigue (col 12-13, rows 100-105) -
    fatigue = {
        "0-20":   int(len(df[df["llm_seller_fatigue_probability"] < 0.20])),
        "20-40":  int(len(df[(df["llm_seller_fatigue_probability"] >= 0.20) & (df["llm_seller_fatigue_probability"] < 0.40)])),
        "40-60":  int(len(df[(df["llm_seller_fatigue_probability"] >= 0.40) & (df["llm_seller_fatigue_probability"] < 0.60)])),
        "60-80":  int(len(df[(df["llm_seller_fatigue_probability"] >= 0.60) & (df["llm_seller_fatigue_probability"] < 0.80)])),
        "80-100": int(len(df[df["llm_seller_fatigue_probability"] >= 0.80])),
    }
    data_ws.cell(1, 10, "Fatigue"); data_ws.cell(1, 11, "Count")
    for i, (bucket, count) in enumerate(fatigue.items()):
        data_ws.cell(2 + i, 10, bucket)
        data_ws.cell(2 + i, 11, count)
        
def _color_series(series, colors):
    """Apply solid fill colors to each data point in a series."""
    for i, hex_color in enumerate(colors):
        pt = DataPoint(idx=i, invertIfNegative=False)
        pt.graphicalProperties.solidFill = hex_color
        series.dPt.append(pt)

def _add_opportunity_donut(ws, data_ws):

    chart = DoughnutChart()
    chart.holeSize = 55
    chart.firstSliceAng = 90
    chart.width = 10
    chart.height = 4
    chart.legend.position = "r"

    labels = Reference(
        data_ws,
        min_col=1,
        min_row=101,
        max_row=105
    )

    data = Reference(
        data_ws,
        min_col=2,
        min_row=101,
        max_row=105
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    dLbls = DataLabelList()
    dLbls.showVal = False
    dLbls.showCatName = False
    dLbls.showSerName = False
    dLbls.showPercent = False
    chart.series[0].dLbls = dLbls

    _color_series(chart.series[0], DIST_COLORS)
    ws.add_chart(chart, "A12")

def _add_distress_bar(ws, data_ws):
    chart = BarChart()
    chart.type     = "bar"
    chart.grouping = "clustered"
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.y_axis.spPr = None
    chart.x_axis.spPr = None
    
    chart.width    = 10
    chart.height   = 4
    # chart.legend.position = "r"
    chart.legend = None


    data   = Reference(data_ws, min_col=5, min_row=101, max_row=105)
    labels = Reference(data_ws, min_col=4, min_row=101, max_row=105)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    _color_series(chart.series[0], SIG_COLORS)

    dLbls = DataLabelList()
    dLbls.showVal     = True
    dLbls.numFmt = '0"%"'
    dLbls.showLegendKey = False
    dLbls.showCatName = False
    dLbls.showSerName = False
    chart.series[0].dLbls = dLbls

    ws.add_chart(chart, "F12")

def _add_seller_fatigue_chart(dashboard_ws, data_ws):
    chart = LineChart()
    chart.height = 4
    chart.width = 6.45
    chart.legend = None
    chart.smooth = True
    chart.style = 10

    # Y axis
    chart.y_axis.title = "Hotel Count"
    chart.y_axis.axPos = "r"
    chart.y_axis.majorGridlines = None
    chart.y_axis.delete = False

    # X axis
    chart.x_axis.title = None
    chart.x_axis.axPos = "t"
    chart.x_axis.delete = False
 
    data   = Reference(data_ws, min_col=11, min_row=2, max_row=6)
    labels = Reference(data_ws, min_col=10, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    dashboard_ws.add_chart(chart, "J12")

def _add_markets_bar(ws, df, data_ws):
    chart = BarChart()
    chart.type     = "bar"
    chart.grouping = "clustered"
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.y_axis.spPr = None
    chart.x_axis.spPr = None
    chart.y_axis.reverseOrder=False
    
    chart.width    = 10
    chart.height   = 4
    chart.legend = None

    n_markets = min(5, len(df["city"].unique())) if "city" in df.columns else 5
    data   = Reference(data_ws, min_col=8, min_row=101, max_row=100 + n_markets)
    labels = Reference(data_ws, min_col=7, min_row=101, max_row=100 + n_markets)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    _color_series(chart.series[0], MKT_COLORS)

    dLbls = DataLabelList()
    dLbls.showLegendKey = False
    dLbls.showVal     = True
    dLbls.showSerName = False
    dLbls.showCatName = False
    chart.series[0].dLbls = dLbls

    ws.add_chart(chart, "M12")


# ── XML post-patch: hide axis tick number labels ──────────────────────────────
def patch_axis_tick_labels(filepath):
    """
    openpyxl cannot set tickLblPos='none' via API (not in its allowed values).
    We patch the saved xlsx ZIP directly — only line chart axes are touched.
    Axis titles ("Hotel Count", "Seller Fatigue %") are preserved.
    """
    files = {}
    with zipfile.ZipFile(filepath, "r") as zin:
        for name in zin.namelist():
            files[name] = zin.read(name)
 
    for name in list(files.keys()):
        if re.match(r"xl/charts/chart\d+\.xml$", name):
            xml = files[name].decode("utf-8")
            if "<lineChart>" in xml:
                xml = xml.replace("</catAx>", '<tickLblPos val="none"/></catAx>')
                xml = xml.replace("</valAx>",  '<tickLblPos val="none"/></valAx>')
                files[name] = xml.encode("utf-8")
 
    with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)

# ── Main entry point ──────────────────────────────────────────────────────────
def _charts_section(ws, data_ws, df):
    _write_chart_data(data_ws, df)
    title_fill = PatternFill("solid", fgColor="0B5D4A")
    title_font = Font(color="FFFFFF", bold=True, size=10)
    title_align = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("A9:P9")
    ws["A9"] = ""
    # Opportunity
    ws.merge_cells("A10:E10")
    ws["A10"] = "Opportunity Score Distribution"
    
    # Distress
    ws.merge_cells("F10:I10")
    ws["F10"] = "Top Distress Signals (%)"
    
    # Seller Fatigue
    ws.merge_cells("J10:L10")
    ws["J10"] = "Seller Fatigue Probability (%)"
    
    # Markets
    ws.merge_cells("M10:P10")
    ws["M10"] = "Highest Opportunity Markets"

    ws.merge_cells("A11:P11")
    ws["A11"] = ""
    

    for cell_ref in ["A10", "F10", "J10", "M10"]:
        ws[cell_ref].fill = title_fill
        ws[cell_ref].font = title_font
        ws[cell_ref].alignment = title_align
    
    ws.row_dimensions[10].height = 20
    _add_opportunity_donut(ws, data_ws)
    _add_distress_bar(ws, data_ws)
    _add_seller_fatigue_chart(ws, data_ws)
    _add_markets_bar(ws, df, data_ws)
