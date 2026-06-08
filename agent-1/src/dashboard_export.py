# # agent-1/src/dashboard_export.py
from src.dashboard.dashboard_data import *
from src.dashboard.dashboard_metrics import *
from src.dashboard.dashboard_components import *
from src.dashboard.dashboard_sheets import *
from src.dashboard.dashboard_charts import *
from openpyxl import Workbook
from datetime import datetime

def export_dashboard():
    df      = load_leads()
    df      = prepare_leads(df)
    metrics = build_metrics(df)

    wb = Workbook()
    dashboard_ws = wb.active; dashboard_ws.title = "Dashboard"
    data_ws      = wb.create_sheet("_ChartData") 
    all_leads_ws = wb.create_sheet("All Leads")
    high_op_ws   = wb.create_sheet("High Opportunity")
    own_ws       = wb.create_sheet("Ownership Analysis")
    dist_ws      = wb.create_sheet("Distress Signals")
    track_ws     = wb.create_sheet("Lead Tracker")
    cmbs_ws      = wb.create_sheet("CMBS Watchlist")

    data_ws.sheet_state = "hidden"
    build_dashboard(dashboard_ws, data_ws, metrics, df)
    populate_all_leads(all_leads_ws, df)
    populate_high_opportunity(high_op_ws, df)
    populate_ownership_analysis(own_ws, df)
    populate_cmbs_sheet(cmbs_ws, df)
    populate_distress_sheet(dist_ws, df)
    populate_lead_tracker(track_ws, df)

    dashboard_ws.sheet_properties.tabColor = "1A2744"
    all_leads_ws.sheet_properties.tabColor = "0B5D4A"
    track_ws.sheet_properties.tabColor     = "0B5D4A"
    high_op_ws.sheet_properties.tabColor   = "4C1D95"
    cmbs_ws.sheet_properties.tabColor      = "991B1B"
    dist_ws.sheet_properties.tabColor      = "9A3412"

    filename = f"Hotel_Acquisition_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    patch_axis_tick_labels(filename) 
    print(f"✅ Saved: {filename}")
    return filename


if __name__ == "__main__":
    export_dashboard()
