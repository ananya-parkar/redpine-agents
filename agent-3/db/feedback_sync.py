"""
Feedback Sync — reads the dashboard Excel file BEFORE it gets
regenerated this run, pulls out whatever Feedback/Notes a human has
entered on the Top_Companies sheet, and writes those into Postgres.

Must run BEFORE generate_dashboard() in main.py, or the file gets
overwritten with a fresh blank-feedback version first and there's
nothing left to read.

Safety rule: only updates a candidate's status in Postgres if:
    1. The Excel row has a non-blank Feedback value, AND
    2. Postgres still shows that candidate as "New"

This means a human's decision, once recorded, can never be silently
reverted by a stale or re-opened Excel file - the only way to change
a status after the first sync is a deliberate new dropdown selection,
and even that requires the corresponding code path to treat it as an
explicit change (not built here yet - this version only captures the
FIRST decision per candidate).

------------------------------------------------------------------------
LEARNING LOOP (tuning triggers)
------------------------------------------------------------------------
Per the client's requirement: "Bad Data" and "Passed" tags should
reduce false positives over time, and 3+ leads sharing the same root
cause should trigger a fix in the NEXT run - not just a note for a
developer to eventually act on.

How this actually works:

1. Cross-run, cross-search: every time feedback is synced, this checks
   ALL "Bad Data" and "Passed" rows in review_status (any search_
   request_id, any run) for notes that repeat 3+ times (case/whitespace
   -insensitive exact match).

2. Each qualifying pattern is upserted (not duplicated) into
   tuning_triggers as one 'active' row per (root_cause, feedback_type)
   - re-triggering the same pattern updates occurrences/affected
   candidates on the existing row rather than creating a new one each
   run. This keeps the "currently active" list clean for step 3.

3. get_active_learned_instructions_text() pulls every 'active' trigger
   and formats the client's OWN flagged note text (not a guessed
   interpretation of it) into a short instruction block. main.py fetches
   this once per pipeline run and passes it into discover_companies()
   and profile_company(), which append it to their prompts as "known
   issues to avoid." This is the actual next-run behavior change -
   Claude sees "3 people flagged X" and is told to avoid X, without a
   developer touching code in between.

4. This is prompt-level nudging, not deterministic - Claude may not
   fully avoid the flagged issue every time. It also does not rewrite
   pipeline code or filtering logic automatically; if the underlying
   fix genuinely needs a code change (e.g. the dedup threshold, a
   filter in main.py), a developer still has to make that change. The
   `recommendation` field (dev-facing, keyword-matched) stays for that
   purpose - it's a hint for the humans, separate from the client-
   facing note text used for prompt injection.

5. A trigger stays 'active' (and keeps being injected) until a
   developer marks it 'resolved' via db/manage_tuning_triggers.py,
   once they've confirmed whatever fix was needed actually landed.
   Without this, injected instructions would accumulate forever even
   after a bug was fixed.
"""

import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import psycopg2
from dotenv import load_dotenv

load_dotenv(override=True)

DB_CONFIG = {
    "host": os.getenv("POSTGRES_HOST", "localhost"),
    "port": os.getenv("POSTGRES_PORT", "5432"),
    "dbname": os.getenv("POSTGRES_DB_AGENT3", "redpine_agent3"),
    "user": os.getenv("POSTGRES_USER", "postgres"),
    "password": os.getenv("POSTGRES_PASSWORD", "postgres"),
}

VALID_STATUSES = {"New", "Pursuing", "Passed", "Bad Data"}

FEEDBACK_PATTERN_THRESHOLD = 3
LEARNED_TYPES = ("Bad Data", "Passed")

# Dev-facing hint only (shown in console / tuning_triggers.txt / the
# `recommendation` DB column) - NOT what gets injected into the LLM
# prompts. Match is substring, case-insensitive, first match wins.
# Tune as real root-cause language comes in from Matthew's notes.
TUNING_RECOMMENDATIONS = {
    "not private":       "Review the Public-company exclusion in discover_companies()'s "
                          "prompt (discovery/company_discovery.py) and the company_type "
                          "check in main.py - public companies may be slipping through.",
    "wrong industry":    "Review industry matching in discover_companies() "
                          "(discovery/company_discovery.py) - industry keyword/criteria "
                          "matching may be too loose.",
    "not family owned":  "Review the family_owned extraction prompt in profile_company() "
                          "(collection/company_profile.py) - may need stronger evidence "
                          "requirements before returning 'Yes'.",
    "not founder led":   "Review the founder_led extraction prompt in profile_company() "
                          "(collection/company_profile.py) - may need stronger evidence "
                          "requirements before returning 'Yes'.",
    "wrong location":    "Review the state-matching logic in main.py (company_state vs "
                          "expected_full/expected_abbr) - geography filter may be too loose.",
    "wrong state":       "Review the state-matching logic in main.py (company_state vs "
                          "expected_full/expected_abbr) - geography filter may be too loose.",
    "duplicate":         "Review deduplication/dedupe.py - fuzzy_match_score threshold or "
                          "normalize_company_name() may need tuning.",
    "too small":         "Review revenue/size signal extraction in profile_company() "
                          "(collection/company_profile.py) - company size may need to "
                          "surface earlier, at the discovery stage.",
    "too big":           "Review revenue/size signal extraction in profile_company() "
                          "(collection/company_profile.py) - company size may need to "
                          "surface earlier, at the discovery stage.",
    "already known":     "Review the 'relatively unknown outside local markets' instruction "
                          "in discover_companies()'s prompt (discovery/company_discovery.py) "
                          "- may need reinforcement.",
    "stale":             "Review the Search Priority ordering in profile_company() "
                          "(collection/company_profile.py) - may need fresher-source "
                          "prioritization.",
    "wrong owner":       "Review founder_name extraction in profile_company() "
                          "(collection/company_profile.py) - leadership-page parsing may be "
                          "picking up the wrong exec.",
    "public company":    "Review the Public-company exclusion in discover_companies()'s "
                          "prompt (discovery/company_discovery.py) and the company_type "
                          "check in main.py - public companies may be slipping through.",
}


def _get_recommendation(root_cause, feedback_type="Bad Data"):
    """Dev-facing hint - match a flagged note to a known fix by substring."""
    rc_lower = root_cause.lower()
    for keyword, rec in TUNING_RECOMMENDATIONS.items():
        if keyword in rc_lower:
            return rec
    if feedback_type == "Passed":
        return (f"Review pipeline for recurring 'Passed' reason: '{root_cause}'. "
                f"This may reflect a soft preference not captured in Search Criteria "
                f"(input/search_request.xlsx) rather than a bug - confirm with the "
                f"client before changing filtering logic.")
    return (f"Review pipeline for root cause: '{root_cause}'. "
            f"Check discover_companies() and profile_company() prompts, "
            f"and the filtering logic in main.py.")


def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def read_feedback_from_excel(dashboard_file_path):
    """
    Returns a list of dicts: [{"candidate_id": int, "feedback": str, "notes": str}, ...]
    Reads only rows with a non-blank Feedback value.
    """
    if not os.path.exists(dashboard_file_path):
        print(f"No existing dashboard file at {dashboard_file_path} - nothing to sync.")
        return []

    wb = load_workbook(dashboard_file_path, data_only=True)
    if "Top_Companies" not in wb.sheetnames:
        print("Top_Companies sheet not found - nothing to sync.")
        return []

    ws = wb["Top_Companies"]

    # Find the relevant columns by header name, rather than hardcoding
    # column letters - keeps this resilient to column reordering.
    header_row = [cell.value for cell in ws[1]]
    try:
        feedback_col = header_row.index("Feedback") + 1
        notes_col = header_row.index("Notes") + 1
        id_col = header_row.index("_candidate_id") + 1
    except ValueError:
        print("Expected columns (Feedback, Notes, _candidate_id) not found "
              "in Top_Companies - nothing to sync. (Is this an older "
              "dashboard file from before feedback columns were added?)")
        return []

    results = []
    for row in ws.iter_rows(min_row=2):
        candidate_id = row[id_col - 1].value
        feedback = row[feedback_col - 1].value
        notes = row[notes_col - 1].value

        if candidate_id is None:
            continue
        if not feedback or str(feedback).strip() == "":
            continue
        feedback = str(feedback).strip()
        if feedback not in VALID_STATUSES:
            print(f"Skipping candidate_id {candidate_id}: "
                  f"unrecognized feedback value '{feedback}'")
            continue

        results.append({
            "candidate_id": int(candidate_id),
            "feedback": feedback,
            "notes": str(notes).strip() if notes else None,
        })

    return results


def apply_feedback_to_postgres(feedback_rows):
    """
    Applies each feedback row to Postgres, but ONLY if the candidate's
    current status in Postgres is still 'New'. This is the safety rule
    that prevents a stale/reopened Excel file from reverting a real
    decision someone already made.

    Returns a stats dict, including how many Bad Data / Passed rows
    were applied THIS run - used by the caller to decide whether it's
    worth checking for learning-loop patterns at all.
    """
    stats = {"applied": 0, "skipped": 0, "bad_data_applied": 0, "passed_applied": 0}

    if not feedback_rows:
        print("No feedback rows to apply.")
        return stats

    conn = get_connection()

    try:
        with conn.cursor() as cur:
            for item in feedback_rows:
                cur.execute(
                    "SELECT status FROM review_status WHERE candidate_id = %s",
                    (item["candidate_id"],),
                )
                row = cur.fetchone()
                current_status = row[0] if row else None

                if current_status != "New":
                    # Already reviewed via some other path - don't touch it.
                    stats["skipped"] += 1
                    continue

                cur.execute(
                    """
                    UPDATE review_status
                    SET status = %s,
                        comments = %s,
                        updated_at = NOW()
                    WHERE candidate_id = %s
                    """,
                    (item["feedback"], item["notes"], item["candidate_id"]),
                )
                stats["applied"] += 1
                if item["feedback"] == "Bad Data":
                    stats["bad_data_applied"] += 1
                elif item["feedback"] == "Passed":
                    stats["passed_applied"] += 1

        conn.commit()
        print(f"Applied feedback to {stats['applied']} of {len(feedback_rows)} "
              f"candidates ({stats['skipped']} already reviewed, so skipped).")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return stats


# ---------------------------------------------------------------------------
# Pattern detection -> tuning triggers (Bad Data AND Passed)
# ---------------------------------------------------------------------------

def get_feedback_patterns(conn, status, threshold=FEEDBACK_PATTERN_THRESHOLD):
    """
    Cross-run, cross-search_request: looks at every candidate ever
    marked with the given status (`Bad Data` or `Passed`) in
    review_status, groups by note text (case/whitespace-insensitive
    exact match), and returns groups at/above `threshold`.

    NOTE: exact-match grouping means "wrong owner" and "wrong owner "
    (trailing space) collapse together (handled via TRIM), but "wrong
    owner" and "incorrect owner" do NOT - they'd count as two separate
    patterns. If notes turn out too free-form to cluster this way in
    practice, this is the place to add fuzzy matching later (dedupe.py
    already uses rapidfuzz for a similar problem on company names).
    """
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT LOWER(TRIM(rs.comments)) AS root_cause,
                   COUNT(*) AS occurrences,
                   STRING_AGG(c.company_name, ', ') AS affected_candidates
            FROM review_status rs
            JOIN candidates c ON c.id = rs.candidate_id
            WHERE rs.status = %s
              AND rs.comments IS NOT NULL
              AND TRIM(rs.comments) != ''
            GROUP BY LOWER(TRIM(rs.comments))
            HAVING COUNT(*) >= %s
            ORDER BY COUNT(*) DESC
            """,
            (status, threshold),
        )
        rows = cur.fetchall()

    return [
        {"root_cause": r[0], "occurrences": r[1], "affected_candidates": r[2]}
        for r in rows
    ]


def upsert_tuning_trigger(conn, root_cause, feedback_type, occurrences,
                           affected_candidates, recommendation):
    """
    One 'active' row per (root_cause, feedback_type). Re-detecting the
    same pattern on a later run UPDATES the existing active row instead
    of inserting a duplicate - keeps the active list (and therefore the
    prompt-injection text) from growing the same line over and over.
    """
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id FROM tuning_triggers
            WHERE root_cause = %s AND feedback_type = %s AND status = 'active'
            """,
            (root_cause, feedback_type),
        )
        existing = cur.fetchone()

        if existing:
            cur.execute(
                """
                UPDATE tuning_triggers
                SET occurrences = %s,
                    affected_candidates = %s,
                    recommendation = %s,
                    updated_at = NOW()
                WHERE id = %s
                """,
                (occurrences, affected_candidates, recommendation, existing[0]),
            )
        else:
            cur.execute(
                """
                INSERT INTO tuning_triggers
                    (root_cause, feedback_type, occurrences,
                     affected_candidates, recommendation, status)
                VALUES (%s, %s, %s, %s, %s, 'active')
                """,
                (root_cause, feedback_type, occurrences,
                 affected_candidates, recommendation),
            )
    conn.commit()


def check_and_log_tuning_triggers():
    """
    Checks for 3+ candidates sharing the same note under 'Bad Data' AND
    under 'Passed', across ALL search requests and ALL runs. Upserts
    each qualifying pattern into tuning_triggers (active), logs a
    dev-facing recommendation to console + tuning_triggers.txt.

    Does NOT rewrite pipeline code. What it DOES do: after this call,
    get_active_learned_instructions_text() will include this pattern,
    so the NEXT discover_companies()/profile_company() call actually
    sees "N people flagged this" - that's the real next-run effect.
    """
    conn = get_connection()
    try:
        any_found = False

        for feedback_type in LEARNED_TYPES:
            patterns = get_feedback_patterns(conn, feedback_type)
            if not patterns:
                continue

            any_found = True
            print(f"\n  [TUNING TRIGGERS DETECTED - {feedback_type}]")
            report_lines = []

            for p in patterns:
                root_cause = p["root_cause"] or "unspecified"
                count = p["occurrences"]
                affected = p["affected_candidates"]
                rec = _get_recommendation(root_cause, feedback_type)

                print(f"\n  Root cause : '{root_cause}'")
                print(f"  Type       : {feedback_type}")
                print(f"  Count      : {count} candidates")
                print(f"  Affected   : {affected[:80]}")
                print(f"  ACTION     : {rec}")

                upsert_tuning_trigger(conn, root_cause, feedback_type, count, affected, rec)

                report_lines.append(
                    f"\nType        : {feedback_type}\n"
                    f"Root cause  : {root_cause}\n"
                    f"Occurrences : {count}\n"
                    f"Affected    : {affected}\n"
                    f"Fix         : {rec}\n"
                )

            report_path = Path(__file__).parent / "tuning_triggers.txt"
            with open(report_path, "a") as f:
                f.write(f"\n{'=' * 60}\n")
                f.write(f"TUNING TRIGGERS - {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
                f.write(f"{'=' * 60}\n")
                f.writelines(report_lines)
            print(f"\n  Tuning report saved -> {report_path}")

        if not any_found:
            print("\n  No feedback pattern has reached the threshold yet.")

    finally:
        conn.close()


def get_active_learned_instructions_text(limit=10):
    """
    Pulls every 'active' tuning trigger and formats the CLIENT'S OWN
    flagged note text (not a guessed interpretation of it) into a short
    instruction block. main.py fetches this once per run and passes it
    into discover_companies() and profile_company() so the prompts
    themselves carry forward what's been flagged - this is the actual
    mechanism behind "next iteration" learning.

    Returns "" if there's nothing active (so callers can skip adding an
    empty section to the prompt).
    """
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT root_cause, feedback_type, occurrences
                FROM tuning_triggers
                WHERE status = 'active'
                ORDER BY updated_at DESC NULLS LAST, created_at DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        return ""

    lines = []
    for root_cause, feedback_type, occurrences in rows:
        lines.append(f'- ({feedback_type}, flagged {occurrences}x): "{root_cause}"')
    return "\n".join(lines)


def sync_feedback_from_dashboard(dashboard_file_path):
    """
    Public entry point. Call this in main.py BEFORE generate_dashboard(),
    so the previous run's saved file (with any human feedback) is read
    before it gets overwritten by a fresh regeneration.
    """
    print("\nSyncing feedback from previous dashboard file...\n")
    feedback_rows = read_feedback_from_excel(dashboard_file_path)
    stats = apply_feedback_to_postgres(feedback_rows)

    if stats["bad_data_applied"] > 0 or stats["passed_applied"] > 0:
        check_and_log_tuning_triggers()

    return stats["applied"]