# dashboard_writer.py
# Your existing dashboard code — no changes except variable names
# are already compatible:
#   signals use: signal_tier, published_at, league, signal_type
#   results use: engagement_action, final_score, venue_name, city, state

from collections import Counter
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

NAVY  = "0E2A47"; WHITE = "FFFFFF"; MGRAY = "64748B"; DGRAY = "374151"
GREEN = "1B5E20"; RED   = "C62828"
RC = 19

def _fill(c): return PatternFill("solid", start_color=c, end_color=c)
def _font(bold=False, color="374151", sz=10, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=sz, italic=italic, name=name)
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border(lt="thin", color="D1D5DB"):
    s = Side(style=lt, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _safe_pct(curr, prev):
    if prev == 0: return 100 if curr > 0 else 0
    return round((curr - prev) / prev * 100)
def _days_ago(val):
    if not val: return 9999
    try:
        dt = datetime.fromisoformat(str(val).replace("Z","+00:00"))
        if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt).days
    except: return 9999
def _date_in_range(val, s, e):
    if not val: return False
    try:
        dt = datetime.fromisoformat(str(val).replace("Z","+00:00"))
        if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
        return s <= dt < e
    except: return False
def _weekly_counts(signals, n=8):
    now = datetime.now(timezone.utc)
    counts, labels = [], []
    for w in range(n, 0, -1):
        s = now - timedelta(weeks=w); e = now - timedelta(weeks=w-1)
        counts.append(sum(1 for sig in signals
                          if _date_in_range(sig.get("published_at") or sig.get("run_at"), s, e)))
        labels.append(f"Wk-{w}")
    return labels, counts
def _dl(val=True, pct=False, cat=False):
    dl = DataLabelList()
    dl.showSerName=False; dl.showLegendKey=False
    dl.showCatName=cat; dl.showVal=val; dl.showPercent=pct
    return dl
def _no_gridlines(chart):
    try: chart.y_axis.majorGridlines = None
    except: pass

def _kpi(ws, col, icon, title, value, pct, accent, bg):
    sc = get_column_letter(col); ec = get_column_letter(col + 2)
    nb = Border()
    ws.merge_cells(f"{sc}5:{ec}5")
    c = ws[f"{sc}5"]; c.value = f"{icon}  {title}"
    c.font = _font(bold=True, color=MGRAY, sz=8)
    c.fill = _fill(bg); c.alignment = _align("left"); c.border = nb
    ws.merge_cells(f"{sc}6:{ec}7")
    c = ws[f"{sc}6"]; c.value = value
    c.font = _font(bold=True, color=accent, sz=24)
    c.fill = _fill(bg); c.alignment = _align(); c.border = nb
    ws.merge_cells(f"{sc}8:{ec}8")
    c = ws[f"{sc}8"]
    c.value = f"{'▲' if pct>=0 else '▼'} {abs(pct)}% vs prior 30 days"
    c.font = _font(color=GREEN if pct>=0 else RED, sz=8)
    c.fill = _fill(bg); c.alignment = _align(); c.border = nb
    for r in range(5, 9):
        for ci in range(col, col+3):
            ws.cell(r, ci).fill = _fill(bg); ws.cell(r, ci).border = nb

def _hdr(ws, row, c1, c2, title, bg=NAVY):
    sc = get_column_letter(c1); ec = get_column_letter(c2)
    ws.merge_cells(f"{sc}{row}:{ec}{row}")
    c = ws[f"{sc}{row}"]; c.value = title
    c.font = _font(bold=True, color=WHITE, sz=9); c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def write_dashboard(ws, signals, results):
    """
    signals: list of dicts from DB (signal_tier, published_at, league, signal_type)
    results: list of dicts from DB (engagement_action, final_score, venue_name, etc.)
    """
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 1
    for ci in range(2, 20):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    for ci in range(21, 35):
        ws.column_dimensions[get_column_letter(ci)].width = 0.3
    for ri in range(1, 65): ws.row_dimensions[ri].height = 14
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 26
    ws.row_dimensions[7].height = 26
    ws.row_dimensions[8].height = 14
    ws.row_dimensions[9].height = 16

    for ri in range(1, 65):
        for ci in range(1, 20):
            ws.cell(ri, ci).fill = _fill(WHITE)

    # ── KPI calculations ──────────────────────────────────────
    recent = [s for s in signals if _days_ago(s.get("published_at")) <= 30]
    prior  = [s for s in signals if 30 < _days_ago(s.get("published_at")) <= 60]
    total_sigs = len(signals)
    tier1_sigs = sum(1 for s in signals if s.get("signal_tier")==1)
    act_now    = sum(1 for r in results if r.get("engagement_action")=="engage_now")
    monitor    = sum(1 for r in results if r.get("engagement_action")=="monitor")
    active     = act_now + monitor
    firms      = len({n.strip().lower()
                      for r in results
                      for n in (r.get("stakeholder_names") or "").split(";")
                      if n.strip() and n.strip().lower() != "none identified"})
    pr_sigs = len(prior)
    pr_t1   = sum(1 for s in prior if s.get("signal_tier")==1)
    run_date     = datetime.now().strftime("%b %d, %Y")
    league_counts = Counter(s.get("league","OTHER") for s in signals)
    top_leagues   = sorted(league_counts.items(), key=lambda x: -x[1])[:8]
    tier_c = [sum(1 for s in signals if s.get("signal_tier")==t) for t in [1,2,3,4]]

    # ── Title + Pills ─────────────────────────────────────────
    ws.merge_cells("B1:E1")
    c = ws["B1"]; c.value = "STADIUM CONSTRUCTION LEAD GEN DASHBOARD"
    c.font = _font(bold=True, color=NAVY, sz=13); c.alignment = _align("left"); c.fill = _fill(WHITE)
    ws.merge_cells("B2:E3")
    c = ws["B2"]; c.value = "Real-time intelligence on stadium, arena & convention center"
    c.font = Font(italic=True, color=MGRAY, size=8, name="Calibri")
    c.alignment = _align("left"); c.fill = _fill(WHITE)
    ws.merge_cells("F1:F3")
    ws["F1"].value = "Data as of:"; ws["F1"].font = _font(color=MGRAY, sz=8)
    ws["F1"].alignment = _align("right"); ws["F1"].fill = _fill(WHITE)
    ws.merge_cells("G1:H3")
    ws["G1"].value = run_date
    ws["G1"].font = _font(bold=True, color="1565C0", sz=11)
    ws["G1"].alignment = _align("left"); ws["G1"].fill = _fill(WHITE)

    PILL_CONFIGS = [
        ("I","NFL",    league_counts.get("NFL",0),                "1D4ED8",WHITE),
        ("J","NCAA",   league_counts.get("NCAA",0),               "16A34A",WHITE),
        ("K","NBA",    league_counts.get("NBA",0),                "7C3AED",WHITE),
        ("L","MLB",    league_counts.get("MLB",0),                "EA580C",WHITE),
        ("M","NHL",    league_counts.get("NHL",0),                "0F766E",WHITE),
        ("N","Tier 1", tier_c[0],                                 "D32F2F",WHITE),
        ("O","Tier 2", tier_c[1],                                 "F57C00",WHITE),
        ("P","Tier 3", tier_c[2],                                 "2E7D32",WHITE),
        ("Q","Tier 4", tier_c[3],                                 "1565C0",WHITE),
        ("R","CONV.",  league_counts.get("Convention Center",0),  "475569",WHITE),
    ]
    for col, lbl, val, bg, fg in PILL_CONFIGS:
        ws.merge_cells(f"{col}1:{col}3")
        c = ws[f"{col}1"]
        c.value = f"{lbl}\n{val}"; c.font = _font(bold=True, color=fg, sz=9)
        c.fill = _fill(bg); c.alignment = _align(wrap=True)
        c.border = _border("thin", "FFFFFF")

    ws.merge_cells("S1:S3")
    ws["S1"].value = "📋  All\nLeads ↗"
    ws["S1"].hyperlink = "#'All Leads'!A1"
    ws["S1"].font  = _font(bold=True, color="1E40AF", sz=8)
    ws["S1"].fill  = _fill("DBEAFE")
    ws["S1"].alignment = _align(wrap=True)
    ws["S1"].border = _border("medium", "93C5FD")

    for ci in range(2, 20): ws.cell(4, ci).fill = _fill("E5E7EB")

    # ── KPI cards ─────────────────────────────────────────────
    kpis = [
        (2,  "📡","TOTAL SIGNALS",    total_sigs,
              _safe_pct(len(recent), pr_sigs),     "1D4ED8","EFF6FF"),
        (5,  "🚩","TIER 1 SIGNALS",   tier1_sigs,
              _safe_pct(sum(1 for s in recent if s.get("signal_tier")==1), pr_t1),
              "16A34A","F0FDF4"),
        (8,  "🎯","ACT NOW LEADS",    act_now,  10, "EA580C","FFF7ED"),
        (11, "📊","MONITOR LEADS",    monitor,   5, "7C3AED","F5F3FF"),
        (14, "🏢","FIRMS IDENTIFIED", firms,    15, "0E2A47","F0F9FF"),
        (17, "🏟","ACTIVE PROJECTS*", active,   17, "0F766E","F0FDF4"),
    ]
    for col, icon, title, value, pct, accent, bg in kpis:
        _kpi(ws, col, icon, title, value, pct, accent, bg)

    ws.merge_cells("N8:R8")
    ws["N8"].value = "* Projects classified as Act Now or Monitor"
    ws["N8"].font  = Font(italic=True, color=MGRAY, size=7, name="Calibri")
    ws["N8"].alignment = _align("right"); ws["N8"].fill = _fill("F0F9FF")

    for ci in range(2, 20): ws.cell(9, ci).fill = _fill("E5E7EB")

    # ── Chart data (hidden columns) ───────────────────────────
    for r in range(1, 12):
        for ci in range(21, 35):
            ws.cell(r, ci).font = Font(color="FFFFFF", size=6)

    for r,(lbl,cnt) in enumerate(zip(["Tier 1","Tier 2","Tier 3","Tier 4"], tier_c), 1):
        ws[f"U{r}"] = lbl; ws[f"V{r}"] = cnt

    eng_d = [("Act Now", act_now), ("Monitor", monitor),
             ("Too Late", sum(1 for r in results
                              if r.get("engagement_action") not in ("engage_now","monitor")))]
    for r,(lbl,cnt) in enumerate(eng_d, 1):
        ws[f"X{r}"] = lbl; ws[f"Y{r}"] = cnt

    for r,(lg,cnt) in enumerate(top_leagues, 1):
        ws[f"AA{r}"] = lg.replace("Convention Center","CONV.")
        ws[f"AB{r}"] = cnt

    wk_lbl, wk_cnt = _weekly_counts(signals, 8)
    for r,(lbl,cnt) in enumerate(zip(wk_lbl, wk_cnt), 1):
        ws[f"AD{r}"] = lbl; ws[f"AE{r}"] = cnt

    src = Counter(s.get("signal_type","news") for s in signals)
    ws["AG1"] = "News";       ws["AH1"] = src.get("news", 0)
    ws["AG2"] = "Government"; ws["AH2"] = src.get("government", 0)

    # ── 4 Charts ──────────────────────────────────────────────
    CW = 9; CH = 6
    c1 = BarChart(); c1.type = "col"; c1.grouping = "clustered"
    c1.title = "SIGNALS BY TIER"; c1.width = CW; c1.height = CH
    c1.add_data(Reference(ws, min_col=22, min_row=1, max_row=4), titles_from_data=False)
    c1.set_categories(Reference(ws, min_col=21, min_row=1, max_row=4))
    c1.dataLabels = _dl(val=True); c1.legend = None
    c1.x_axis.delete = False; _no_gridlines(c1)
    try: c1.series[0].graphicalProperties.solidFill = "2563EB"
    except: pass
    ws.add_chart(c1, "B10")

    c2 = DoughnutChart()
    c2.title = "LEADS BY ENGAGEMENT ACTION"; c2.width = CW; c2.height = CH
    c2.add_data(Reference(ws, min_col=25, min_row=1, max_row=3), titles_from_data=False)
    c2.set_categories(Reference(ws, min_col=24, min_row=1, max_row=3))
    c2.dataLabels = _dl(val=True, cat=False)
    ws.add_chart(c2, "F10")

    c3 = BarChart(); c3.type = "col"; c3.grouping = "clustered"
    c3.title = "SIGNALS BY LEAGUE"; c3.width = CW; c3.height = CH
    c3.add_data(Reference(ws, min_col=28, min_row=1, max_row=len(top_leagues)), titles_from_data=False)
    c3.set_categories(Reference(ws, min_col=27, min_row=1, max_row=len(top_leagues)))
    c3.dataLabels = _dl(val=True); c3.legend = None
    c3.x_axis.delete = False; _no_gridlines(c3)
    try: c3.series[0].graphicalProperties.solidFill = "16A34A"
    except: pass
    ws.add_chart(c3, "K10")

    c4 = LineChart()
    c4.title = "SIGNALS TREND (LAST 8 WEEKS)"; c4.width = CW; c4.height = CH
    c4.add_data(Reference(ws, min_col=31, min_row=1, max_row=8), titles_from_data=False)
    c4.set_categories(Reference(ws, min_col=30, min_row=1, max_row=8))
    c4.dataLabels = _dl(val=True); c4.legend = None
    c4.x_axis.delete = False; _no_gridlines(c4)
    try:
        c4.series[0].graphicalProperties.line.solidFill = "1D4ED8"
        c4.series[0].graphicalProperties.line.width = 25000
    except: pass
    ws.add_chart(c4, "P10")

    # ── Tables ────────────────────────────────────────────────
    SEP = 24
    ws.row_dimensions[SEP].height = 4
    for ci in range(2, 20): ws.cell(SEP, ci).fill = _fill("E5E7EB")

    T = SEP + 1; H = T + 1
    TABLE_ROWS = 15

    # Top 10 Act Now
    _hdr(ws, T, 2, 6, "TOP 10 ACT NOW LEADS", "0E2A47")
    ws.row_dimensions[H].height = 16
    for col, h in [(2,"#"),(3,"Venue / Project"),(4,"City, State"),(5,"Tier"),(6,"Score")]:
        c = ws.cell(H, col, h)
        c.font=_font(bold=True,color=WHITE,sz=8)
        c.fill=_fill("0E2A47"); c.alignment=_align(); c.border=_border("thin","4B5563")
    ws.column_dimensions["C"].width = 20

    top_act = [r for r in results if r.get("engagement_action")=="engage_now"][:TABLE_ROWS]
    for i in range(1, TABLE_ROWS + 1):
        row = H+i; bg = "F0FDF4" if i%2 else WHITE
        if i <= len(top_act):
            r = top_act[i-1]
            # final_score from DB, score from in-memory
            sc = r.get("final_score") or r.get("score","")
            vals = [(2,i),(3,(r.get("venue_name") or "")[:22]),
                    (4,f"{r.get('city','')}, {r.get('state','')}"),
                    (5,r.get("signal_tier","")),(6,sc)]
        else:
            vals = [(2,""),(3,""),(4,""),(5,""),(6,"")]
        for col, val in vals:
            c=ws.cell(row,col,val); c.border=_border(); c.alignment=_align("left" if col==3 else "center")
            if col==6 and val:
                try:
                    sc_ = float(val)
                    sbg = ("C8E6C9" if sc_>=85 else "E8F5E9" if sc_>=75 else "FFF3CD" if sc_>=60 else "FFCDD2")
                    c.fill=_fill(sbg); c.font=_font(bold=True,color="1B5E20",sz=9)
                except: c.fill=_fill(bg); c.font=_font(sz=9)
            else:
                c.fill=_fill(bg); c.font=_font(sz=9)
        ws.row_dimensions[row].height = 13

    # All Leads top 15
    _hdr(ws, T, 7, 13, "ALL LEADS (TOP 15)", NAVY)
    for i, h in enumerate(["#","Venue / Project","League","Tier","Engagement","Score","Date"],0):
        c = ws.cell(H, 7+i, h)
        c.font=_font(bold=True,color=WHITE,sz=8)
        c.fill=_fill(NAVY); c.alignment=_align(); c.border=_border("thin","4B5563")

    ENG_STYLE = {
        "engage_now":  ("D1FAE5","065F46","Act Now"),
        "monitor":     ("FEF9C3","854D0E","Monitor"),
        "too_late":    ("FEE2E2","991B1B","Too Late"),
        "insufficient_signal": ("F3F4F6","6B7280","Low Signal"),
    }
    for i, r in enumerate(results[:TABLE_ROWS], 1):
        row=H+i; bg="F9FAFB" if i%2 else WHITE
        eng=r.get("engagement_action","")
        ebg,efg,elbl=ENG_STYLE.get(eng,("F3F4F6","6B7280",eng))
        sc  = r.get("final_score") or r.get("score","")
        pub = r.get("first_detected_at") or r.get("published_at","") or ""
        dt  = str(pub)[:10] if pub else ""
        for col,val in [(7,i),(8,(r.get("venue_name") or "")[:24]),
                        (9,r.get("league","")),(10,r.get("signal_tier","")),
                        (11,elbl),(12,sc),(13,dt)]:
            c=ws.cell(row,col,val); c.border=_border()
            c.alignment=_align("left" if col==8 else "center")
            if col==11: c.font=_font(bold=True,color=efg,sz=8); c.fill=_fill(ebg)
            else: c.font=_font(sz=9); c.fill=_fill(bg)
        ws.row_dimensions[row].height = 13

    # Additional Leads (ranks 16-23 — the leads not already shown in the
    # "All Leads (Top 15)" table above). Replaces the old "New Leads Today"
    # table, which relied on first_detected_at == today and was often
    # empty (most pipeline runs update existing leads rather than insert
    # brand-new ones, so "today's new" rarely had 8 rows to show).
    # `results` is already pre-sorted by tier/score (same order the Top 15
    # table above uses), so this is simply the next slice after it —
    # genuinely new information on the dashboard, not a duplicate view.
    TIER_BG = {1:"C8E6C9",2:"BBDEFB",3:"FFE0B2",4:"FFCDD2"}
    TIER_FG = {1:"1B5E20",2:"0D47A1",3:"BF360C",4:"B71C1C"}
    _hdr(ws, T, 14, RC, "ADDITIONAL LEADS (16-23)", "0E2A47")
    additional = results[TABLE_ROWS:TABLE_ROWS + 8]
    for i, h in enumerate(["#","Venue / Project","League","Tier","Engagement","Score"], 0):
        c = ws.cell(H, 14+i, h)
        c.font=_font(bold=True,color=WHITE,sz=8)
        c.fill=_fill("0E2A47"); c.alignment=_align(); c.border=_border("thin","4B5563")
    ws.column_dimensions["O"].width = 18
    ADD_ROWS = 8
    for i in range(1, ADD_ROWS + 1):
        row = H+i; bg = "EFF6FF" if i%2 else WHITE
        if i <= len(additional):
            r = additional[i-1]
            sc  = r.get("final_score") or r.get("score","")
            t   = r.get("signal_tier")
            eng = r.get("engagement_action","")
            ebg,efg,elbl = ENG_STYLE.get(eng,("F3F4F6","6B7280",eng))
            vals = [(14,i),(15,(r.get("venue_name") or "")[:22]),
                    (16,r.get("league","")),(17,t),(18,elbl),(19,sc)]
        else:
            t = None; ebg=efg=elbl=None
            vals = [(14,""),(15,""),(16,""),(17,""),(18,""),(19,"")]
        for col, val in vals:
            c=ws.cell(row,col,val); c.border=_border()
            c.alignment=_align("left" if col==15 else "center")
            if col==17 and t and t in TIER_BG:
                c.fill=_fill(TIER_BG[t]); c.font=_font(bold=True,color=TIER_FG[t],sz=9)
            elif col==18 and elbl:
                c.fill=_fill(ebg); c.font=_font(bold=True,color=efg,sz=8)
            else:
                c.fill=_fill(bg); c.font=_font(sz=9)
        ws.row_dimensions[row].height = 13
    if not additional:
        ws.cell(H+1, 14, "No additional leads beyond the Top 15").font = _font(italic=True, color=MGRAY, sz=9)

    # Key Insights
    KEY = T + 11
    _hdr(ws, KEY, 14, RC, "KEY INSIGHTS (LAST 30 DAYS)", "0E2A47")
    rc_t1 = sum(1 for s in recent if s.get("signal_tier")==1)
    top_n  = top_leagues[0][0] if top_leagues else "N/A"
    top_c  = top_leagues[0][1] if top_leagues else 0
    insights = [
        f"Tier 1 signals: {rc_t1} detected in last 30 days.",
        f"{act_now} high-priority (Act Now) leads identified.",
        f"{top_n} has the highest signal volume ({top_c}).",
        f"{active} active projects (Act Now + Monitor) tracked.",
        f"{firms} unique firms identified across all leads.",
    ]
    ec_str = get_column_letter(RC)
    for i, text in enumerate(insights):
        row = KEY+1+i
        ws.merge_cells(f"N{row}:{ec_str}{row}")
        c = ws[f"N{row}"]; c.value = f"✓  {text}"
        c.font=_font(color=DGRAY,sz=9); c.fill=_fill("F5F3FF")
        c.alignment=_align("left",wrap=True); c.border=_border("thin","DDD6FE")
        ws.row_dimensions[row].height = 18