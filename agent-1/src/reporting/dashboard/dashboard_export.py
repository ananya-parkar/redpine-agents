# agent-1/src/reporting/dashboard/dashboard_export.py
from src.reporting.dashboard.dashboard_data import *
from src.reporting.dashboard.dashboard_metrics import *
from src.reporting.dashboard.dashboard_components import *
from src.reporting.dashboard.dashboard_sheets import *
from src.reporting.dashboard.dashboard_charts import *
from openpyxl import Workbook
from datetime import datetime

def export_dashboard(output_dir, run_rows, search_area=""):
    df      = pd.DataFrame(run_rows)
    df      = prepare_leads(df)
    # search_area = ""

    # try:
    #     with open("inputs/locations.txt", "r") as f:
    #         locations = [line.strip() for line in f.readlines() if line.strip()]
    #         search_area = ", ".join(locations)
    # except:
    #     search_area = "Unknown"
   
    print("\n===== DASHBOARD COLUMNS =====")
    # print(df.columns.tolist())
    print(df["owner_name"].notna().sum())
    print(df["room_count"].head(20))
    metrics = build_metrics(df, search_area=search_area)

    wb = Workbook()
    dashboard_ws = wb.active; dashboard_ws.title = "Dashboard"
    data_ws      = wb.create_sheet("_ChartData") 
    all_leads_ws = wb.create_sheet("All Leads")
    high_op_ws   = wb.create_sheet("High Opportunity")
    own_ws       = wb.create_sheet("Ownership Analysis")
    dist_ws      = wb.create_sheet("Distress Signals")
    track_ws     = wb.create_sheet("Lead Tracker")
    # cmbs_ws      = wb.create_sheet("CMBS Watchlist")
    feedback_ws  = wb.create_sheet("Feedback Learning")

    data_ws.sheet_state = "hidden"
    build_dashboard(dashboard_ws, data_ws, metrics, df)
    # print(df.columns.tolist())
    populate_all_leads(all_leads_ws, df)
    populate_high_opportunity(high_op_ws, df)
    populate_ownership_analysis(own_ws, df)
    # populate_cmbs_sheet(cmbs_ws, df)
    populate_distress_sheet(dist_ws, df)
    populate_lead_tracker(track_ws, df)
    populate_feedback_learning(feedback_ws)

    dashboard_ws.sheet_properties.tabColor = "1A2744"
    all_leads_ws.sheet_properties.tabColor = "0B5D4A"
    track_ws.sheet_properties.tabColor     = "0B5D4A"
    high_op_ws.sheet_properties.tabColor   = "4C1D95"
    # cmbs_ws.sheet_properties.tabColor      = "991B1B"
    dist_ws.sheet_properties.tabColor      = "9A3412"

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = (output_dir / f"Hotel_Acquisition_Dashboard_{timestamp}.xlsx")
    wb.save(filename)
    patch_axis_tick_labels(filename) 
    print(f"✅ Saved: {filename}")
    return filename


if __name__ == "__main__":
    export_dashboard()
