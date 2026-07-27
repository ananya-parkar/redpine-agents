# agent-3/dashboard/dashboard.py
import os
import re
from datetime import datetime, timedelta
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, DoughnutChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties

load_dotenv(override=True)
 
DB_CONFIG = {
    "host": os.getenv("POSTGRES_HOST", "localhost"),
    "port": os.getenv("POSTGRES_PORT", "5432"),
    "dbname": os.getenv("POSTGRES_DB_AGENT3", "redpine_agent3"),
    "user": os.getenv("POSTGRES_USER", "postgres"),
    "password": os.getenv("POSTGRES_PASSWORD", "postgres"),
}
 
FONT_NAME = "Calibri"
NAVY = "0B1F48"
NAVY_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
WHITE_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=20)
WHITE_SUBFONT = Font(name=FONT_NAME, color="DCE6F5", italic=True, size=10)
WHITE_SMALL = Font(name=FONT_NAME, color="FFFFFF", size=12)
INVISIBLE_FONT = Font(name=FONT_NAME, size=9, color="FFFFFF")
CURRENT_SEARCH_FILL = PatternFill(
    "solid",
    start_color="F8F9FB",   # very light grey
    end_color="F8F9FB"
)

CARD_BLUE = PatternFill("solid","EDF5FF","EDF5FF")
CARD_GREEN = PatternFill("solid","F3FBEE","F3FBEE")
CARD_YELLOW = PatternFill("solid","FFF9E8","FFF9E8")
CARD_PURPLE = PatternFill("solid","F3EDF9","F3EDF9")
CARD_LIGHTBLUE = PatternFill("solid","EEF7FF","EEF7FF")

HEADER_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
SECTION_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12, color="1B2A4A")
THIN_BORDER = Border(bottom=Side(style="hair", color="ECECEC"))

GREEN_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
RED_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
INSIGHT_FILL = PatternFill(
    "solid",
    start_color="F8FAF3",
    end_color="F8FAF3"
)

STATUS_COLORS = {
    "New": "1F7A1F",
    "Pursuing": "1565C0",
    "Passed": "9E9E9E",
    "Bad Data": "C62828",
}
 
TOTAL_COLS = 20
 
 
def get_connection():
    return psycopg2.connect(**DB_CONFIG)
 
 
def fetch_all_candidates(search_request_id):
    """
    Scoped to the CURRENT search request. If the client switches
    Florida -> Texas, this returns only the Texas leads.
    """
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    c.id AS candidate_id,
                    c.company_name, c.website, c.city, c.state, c.industry,
                    c.company_type, c.company_description,
                    c.founded_year, c.years_in_business, c.founder_name,
                    c.founder_led, c.family_owned, c.founder_age_estimate,
                    c.ownership_status, c.ownership_tenure_years,
                    c.seller_readiness_score, c.fit_analysis,
                    c.seller_readiness_signals, c.why_discovered,
                    rs.status AS review_status,
                    rs.comments AS review_notes,
                    c.first_seen_date, c.last_seen_date,
                    e.why_selected, e.evidence_summary, e.one_line_reason, e.raw_evidence
                FROM candidates c
                LEFT JOIN review_status rs ON rs.candidate_id = c.id
                LEFT JOIN LATERAL (
                    SELECT * FROM evidence ev WHERE ev.candidate_id = c.id
                    ORDER BY ev.created_at DESC LIMIT 1
                ) e ON true
                WHERE c.search_request_id = %s
                ORDER BY c.first_seen_date DESC NULLS LAST, c.company_name ASC
                """,
                (search_request_id,),
            )
            return cur.fetchall()
    finally:
        conn.close()
 
 
def fetch_last_week_snapshot(search_request_id):
    """
    Also scoped - otherwise "vs Last Week" would compare this week's
    Texas run against last week's Florida run.
    """
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT to_regclass('public.pipeline_runs') IS NOT NULL AS exists")
            if not cur.fetchone()["exists"]:
                return None
            target_date = datetime.now().date() - timedelta(days=7)
            cur.execute(
                """
                SELECT * FROM pipeline_runs
                WHERE search_request_id = %s
                ORDER BY ABS(run_date - %s::date) ASC
                LIMIT 1
                """,
                (search_request_id, target_date),
            )
            return cur.fetchone()
    finally:
        conn.close()
 
 
def autosize_columns(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
 
 
def score_fill(score):
    if score is None:
        return None
    if score >= 70:
        return GREEN_FILL
    if score >= 40:
        return YELLOW_FILL
    return RED_FILL
 
 
def status_font(status):
    color = STATUS_COLORS.get(status, "424242")
    return Font(name=FONT_NAME, size=10, bold=True, color=color)
 
 
def delta_text(current, previous):
    if previous is None:
        return ""
    diff = current - previous
    if diff > 0:
        return f"\u2191 {diff}"
    if diff < 0:
        return f"\u2193 {abs(diff)}"
    return "\u2014 0"
 
 
def style_chart(chart):
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.y_axis.delete = True
    chart.x_axis.delete = False
    chart.graphical_properties = None
 
 
# ---------------------------------------------------------------------------
# Display-layer state normalization. NOTE: this is a patch, not the real
# fix - the real fix belongs upstream in signal_extractor.py. Keep this
# even after that lands; it's harmless on already-clean input and
# protects the dashboard from any inconsistency that slips through.
# ---------------------------------------------------------------------------
US_STATE_ABBREV_TO_NAME = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming",
}
 
 
def normalize_state_name(state):
    """
    Collapses abbreviation/full-name inconsistencies (e.g. "FL" and
    "Florida") so they group together in charts and counts instead of
    appearing as separate entries. Returns "Unknown" for blank/missing.
    """
    if not state or not str(state).strip():
        return "Unknown"
    state = str(state).strip()
    if state.upper() in US_STATE_ABBREV_TO_NAME:
        return US_STATE_ABBREV_TO_NAME[state.upper()]
    return state.title()
 
 
def cluster_similar_labels(labels, threshold=75):
    """
    Groups near-duplicate free-text labels into clusters instead of
    counting exact strings only. Needed specifically for `industry`,
    since that field is written fresh by the LLM per company - it might
    write "Precision Machining" for one company and "Precision CNC
    Machining" for another, even though a human would call both the
    same industry. Exact-match counting (what state/ownership_status
    use, fine there since those are controlled values) would fragment
    this into many count=1 bars instead of one accurate count.
 
    Same technique as fuzzy_match_score() in deduplication/dedupe.py,
    applied here to industry labels rather than company names.
 
    Returns [(representative_label, count), ...] - the representative
    is whichever original label was seen first in that cluster. Not
    sorted; caller sorts as needed.
    """
    clusters = []  # list of [representative_label, count]
    for label in labels:
        matched = False
        for cluster in clusters:
            if fuzz.token_sort_ratio(label, cluster[0]) >= threshold:
                cluster[1] += 1
                matched = True
                break
        if not matched:
            clusters.append([label, 1])
    return [(label, count) for label, count in clusters]
 
 
def industry_primary_segment(label):
    """
    Industry values are LLM free text formatted like "Category /
    Sub-category (details)" - e.g. "Precision CNC Machining /
    Aerospace & Defense Manufacturing". The sub-category and
    parenthetical detail add enough noise that fuzzy-matching the FULL
    string barely ever clears a sane threshold (~64 for two things a
    human would call the same industry) even with cluster_similar_
    labels(). Comparing just the primary segment - before the first
    "/" or "(" - fixes that: the same real industries then score
    ~90-100 instead of ~64. Also makes for a cleaner chart label than
    the full noisy string.
    """
    if not label:
        return label
    return re.split(r"[/(]", label)[0].strip()

def fetch_search_request(search_request_id):
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT *
                FROM search_requests
                WHERE id=%s
            """,(search_request_id,))
            return cur.fetchone()
    finally:
        conn.close()

def build_dashboard_sheet(wb, candidates, last_week, search_request):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = "Agent 3 - Business Acquisition Dashboard"
    title_cell.font = WHITE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("A3:G3")
    sub_cell = ws["A3"]
    sub_cell.value = "Surface private, founder-led and family-owned businesses that match the acquisition profile."
    sub_cell.font = WHITE_SUBFONT
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Data as of label
    ws["H1"] = "Data as of:"
    ws["H1"].font = Font(
        name=FONT_NAME,
        size=10,
        color="DCE6F5",
        italic=True
    )
    ws["H1"].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )

    # Timestamp
    ws.merge_cells("I1:J1")

    date_value = ws["I1"]
    date_value.value = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    date_value.font = Font(
        name=FONT_NAME,
        size=11,
        bold=True,
        color="FFFFFF"
    )
    date_value.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws["H2"] = "Run ID:"
    ws["H2"].font = Font(
        name=FONT_NAME,
        size=10,
        color="DCE6F5",
        italic=True
    )
    ws["H2"].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )

    ws.merge_cells("I2:J2")

    run_value = ws["I2"]
    run_value.value = f"RUN-{datetime.now().strftime('%Y%m%d-%H%M')}"
    run_value.font = Font(
        name=FONT_NAME,
        size=10,
        color="FFFFFF",
        bold=True
    )
    run_value.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for row in range(1,5):
        for col in range(1,12):      # A:J
            ws.cell(row, col).fill = NAVY_FILL
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 12
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 20
        ws.row_dimensions[7].height = 18
        ws.row_dimensions[8].height = 12
        ws.row_dimensions[9].height = 20
        ws.row_dimensions[11].height = 16

    total_targets = len(candidates)
    new_this_week = sum(1 for c in candidates if c.get("first_seen_date") == datetime.now().date())
    shortlisted = sum(1 for c in candidates if c.get("review_status") == "Pursuing")
    in_review = sum(1 for c in candidates if c.get("review_status") == "New")
    reviewed = sum(1 for c in candidates if c.get("review_status") in ("Pursuing", "Passed", "Bad Data"))

    founder_led_count = sum(1 for c in candidates if c.get("founder_led") == "Yes")
    family_owned_count = sum(1 for c in candidates if c.get("family_owned") == "Yes")
    scored_only = [c.get("seller_readiness_score") for c in candidates if c.get("seller_readiness_score") is not None]
    highest_score = max(scored_only) if scored_only else 0
    private_count = sum(1 for c in candidates if c.get("company_type") == "Private")

    high_scorers = sum(
        1 for c in candidates
        if (c.get("seller_readiness_score") or 0) >= 80
    )

    kpi_defs = [
        ("🎯", "Companies Found", total_targets, "All Time", "B35C6A"),          # Rose
        ("⭐", "High Opportunity", high_scorers, "Score ≥ 80", "D29A18"),        # Mustard
        ("🏢", "Private Companies", private_count, "100% of total", "C97A38"),  # Copper
        ("👥", "Family-Owned", family_owned_count,
        f"{round(family_owned_count*100/total_targets) if total_targets else 0}% of total",
        "5B9A5B"),                                           
        ("👤","Founder-Led",founder_led_count,
        f"{round(founder_led_count*100/total_targets) if total_targets else 0}% of total",
        "8366B8"),                                                         # Purple
        ("➕", "New This Run", new_this_week,
        f"↑ {new_this_week}",
        "149C9C"),    
    ]

    # CURRENT SEARCH

    # Row 5 title
    ws.merge_cells("A5:K5")

    title = ws["A5"]
    title.value = "CURRENT SEARCH"
    title.font = Font(
        name=FONT_NAME,
        bold=True,
        size=13,
        color="1B2A4A"
    )
    title.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    ws.row_dimensions[5].height = 22


    # Background card (Rows 6-8)

    CARD_BORDER = Border(
        left=Side(style="medium", color="FFFFFF"),
        right=Side(style="medium", color="FFFFFF"),
        top=Side(style="medium", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )

    for r in range(6, 9):
        for c in range(1, 12):          # A:K
            cell = ws.cell(r, c)
            cell.fill = CURRENT_SEARCH_FILL
            cell.border = CARD_BORDER

    search_items = [
        ("📍", "Geography", search_request.get("geography") or "-"),
        ("🏭", "Industry", search_request.get("industry") or "-"),
        ("👥", "Ownership", search_request.get("ownership_preference") or "-"),
        ("📅", "Min Years", f"{search_request.get('min_years') or '-'}+ Years"),
        ("👤", "Founder Age", search_request.get("founder_age") or "-"),
    ]

    positions = [1,4,6,8,10]

    for (icon, heading, value), col in zip(search_items, positions):

        merge_widths = {
            1: 3,   # A:C
            4: 2,   # D:E
            6: 2,   # F:G
            8: 2,   # H:I
            10: 2,  # J:K
        }

        ws.merge_cells(
            start_row=6,
            start_column=col,
            end_row=8,
            end_column=col + merge_widths[col] - 1
        )

        cell = ws.cell(row=6, column=col)

        cell.value = (
            f"{icon}\n"
            f"{heading}\n"
            f"{value}"
        )

        cell.font = Font(
            name=FONT_NAME,
            size=11,
            bold=True,
            color="1B2A4A"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="top",
            wrap_text=True
        )

    CARD_BORDER = Border(
        left=Side(style="thin", color="ECECEC"),
        right=Side(style="thin", color="ECECEC"),
        top=Side(style="thin", color="ECECEC"),
        bottom=Side(style="thin", color="ECECEC"),
    )

    CARD_COLS = [
        (1,2),    # A:B
        (3,4),    # C:D
        (5,6),    # E:F
        (7,8),    # G:H
        (9,10),   # I:J
        (11,11),
    ]

    ws.merge_cells("A9:K9")
    c = ws["A9"]
    c.value = "TOP SIGNALS"
    c.font = Font(
        name=FONT_NAME,
        size=12,
        bold=True,
        color="FFFFFF"
    )

    c.fill = NAVY_FILL
    c.alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    # ================= KPI CARDS ====================
    ws.row_dimensions[10].height = 56
    ws.row_dimensions[11].height = 12
    ws.row_dimensions[12].height = 12
    ws.row_dimensions[13].height = 12

    for (icon, title, value, caption, color), (start_col, end_col) in zip(kpi_defs, CARD_COLS):

        fill = PatternFill(
            "solid",
            start_color=color,
            end_color=color
        )

        ws.merge_cells(
            start_row=10,
            start_column=start_col,
            end_row=13,
            end_column=end_col
        )

        cell = ws.cell(row=10, column=start_col)

        cell.value = (
            f"{icon}\n"
            f"{title}\n"
            f"{value}\n"
            f"{caption}"
        )

        cell.fill = fill

        cell.font = Font(
            name=FONT_NAME,
            size=13,
            bold=True,
            color="FFFFFF"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for r in range(10,14):
            for c in range(start_col,end_col+1):
                ws.cell(r,c).fill = fill
                ws.cell(r,c).border = CARD_BORDER

    # ================= CHART CARDS ====================
    section_headers = [
        ("A14:C14", "TOP GEOGRAPHIES"),
        ("D14:F14", "OWNERSHIP MIX"),
        ("G14:I14", "INDUSTRY BREAKDOWN"),
        ("J14:K14", "KEY INSIGHTS"),
    ]

    for rng, title in section_headers:

        ws.merge_cells(rng)

        cell = ws[rng.split(":")[0]]
        cell.value = title
        cell.fill = NAVY_FILL

        cell.font = Font(
            name=FONT_NAME,
            size=12,
            bold=True,
            color="FFFFFF"
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

    ws.row_dimensions[14].height = 22

    CARD_FILL = PatternFill(
        "solid",
        start_color="FCFCFC",
        end_color="FCFCFC"
    )

    CHART_BORDER = Border(
        left=Side(style="thin", color="D8D8D8"),
        right=Side(style="thin", color="D8D8D8"),
        top=Side(style="thin", color="D8D8D8"),
        bottom=Side(style="thin", color="D8D8D8"),
    )

    chart_cards = [
        ("A15:C32"),
        ("D15:F32"),
        ("G15:I32"),
        ("J15:K32"),
    ]

    for rng in chart_cards:

        start, end = rng.split(":")

        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        s_col, s_row = coordinate_from_string(start)
        e_col, e_row = coordinate_from_string(end)

        for r in range(s_row, e_row + 1):
            for c in range(
                column_index_from_string(s_col),
                column_index_from_string(e_col) + 1
            ):
                cell = ws.cell(r, c)
                cell.fill = CARD_FILL
                cell.border = CHART_BORDER

    for r in range(15, 33):
        ws.row_dimensions[r].height = 22

    helper_data_row = 45
    chart_data_row = 15

    state_counts = {}
    for c in candidates:
        st = normalize_state_name(c.get("state"))
        state_counts[st] = state_counts.get(st, 0) + 1
    top_states = sorted(state_counts.items(), key=lambda x: -x[1])[:5]

    geo_header_row = helper_data_row
    ws.cell(row=geo_header_row, column=1, value="State").font = INVISIBLE_FONT
    ws.cell(row=geo_header_row, column=2, value="Count").font = INVISIBLE_FONT
    for i, (state, count) in enumerate(top_states, start=1):
        ws.cell(row=geo_header_row + i, column=1, value=state).font = INVISIBLE_FONT
        ws.cell(row=geo_header_row + i, column=2, value=count).font = INVISIBLE_FONT
    geo_last_row = geo_header_row + len(top_states)
 
    ownership_counts = {}
    for c in candidates:
        status = c.get("ownership_status") or "Unknown"
        ownership_counts[status] = ownership_counts.get(status, 0) + 1
    ownership_items = sorted(ownership_counts.items(), key=lambda x: -x[1])

    ownership_header_row = helper_data_row
    ws.cell(row=ownership_header_row, column=4, value="Ownership").font = INVISIBLE_FONT
    ws.cell(row=ownership_header_row, column=5, value="Count").font = INVISIBLE_FONT
    for i, (status, count) in enumerate(ownership_items, start=1):
        ws.cell(row=ownership_header_row + i, column=4, value=status).font = INVISIBLE_FONT
        ws.cell(row=ownership_header_row + i, column=5, value=count).font = INVISIBLE_FONT
    ownership_last_row = ownership_header_row + len(ownership_items)
 
    # Cluster on the PRIMARY SEGMENT (before "/" or "("), not the full
    # industry string - see industry_primary_segment() docstring for
    # why. Threshold raised to 80 since primary segments are short and
    # clean enough that real matches score ~90-100 here, vs ~64 on the
    # full noisy strings (which is why the chart was collapsing to
    # mostly count=1 bars before this).
    industry_labels = [
        industry_primary_segment(c.get("industry")) or "Unknown" for c in candidates
    ]
    industry_clusters = cluster_similar_labels(industry_labels, threshold=80)
 
    def _truncate_label(label):
        return label[:18] + "..." if len(label) > 18 else label
 
    top_industries = sorted(
        [(_truncate_label(label), count) for label, count in industry_clusters],
        key=lambda x: -x[1],
    )[:4]
 
    # Columns 17-18 (not 7-8) - see comment above cluster_similar_labels
    # usage below. Mini-cards (further down this function) write into
    # columns 8-9 for these exact same rows (11+), so putting the
    # industry chart's hidden source data at column 8 meant the mini-
    # cards silently overwrote the count values with text after the
    # fact - the chart was then plotting text instead of numbers,
    # rendering as blank/zero bars. This does NOT move the chart itself
    # (still visually anchored at F via ws.add_chart below) - only
    # where its underlying data lives.
    industry_header_row = helper_data_row
    ws.cell(row=industry_header_row, column=17, value="Industry").font = INVISIBLE_FONT
    ws.cell(row=industry_header_row, column=18, value="Count").font = INVISIBLE_FONT
    for i, (industry, count) in enumerate(top_industries, start=1):
        ws.cell(row=industry_header_row + i, column=17, value=industry).font = INVISIBLE_FONT
        ws.cell(row=industry_header_row + i, column=18, value=count).font = INVISIBLE_FONT
    industry_last_row = industry_header_row + len(top_industries)
 
    helper_block_last_row = max(geo_last_row, ownership_last_row, industry_last_row)
    for r in range(helper_data_row, helper_block_last_row + 1):
        ws.row_dimensions[r].height = 1

    pie = DoughnutChart()
    pie.holeSize = 58
    pie.legend.position = "r"
    data = Reference(ws, min_col=2, min_row=geo_header_row, max_row=geo_last_row)
    cats = Reference(ws, min_col=1, min_row=geo_header_row + 1, max_row=geo_last_row)
    pie.add_data(data, titles_from_data=True)
    pie.style = 20
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = False
    pie.dataLabels.showCatName = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showVal = False
    pie.dataLabels.showLegendKey = False
    pie.height = 4.5
    pie.width = 9.25
    ws.add_chart(pie, "A15")

    bar1 = BarChart()
    bar1.style = 10
    bar1.legend = None
    bar1.gapWidth = 50
    data = Reference(ws, min_col=5, min_row=ownership_header_row, max_row=ownership_last_row)
    cats = Reference(ws, min_col=4, min_row=ownership_header_row + 1, max_row=ownership_last_row)
    bar1.add_data(data, titles_from_data=True)
    bar1.style = 19
    bar1.set_categories(cats)
    bar1.dataLabels = DataLabelList()
    bar1.dataLabels.showVal = True
    bar1.dataLabels.showCatName = False
    bar1.dataLabels.showSerName = False
    bar1.dataLabels.showLegendKey = False
    style_chart(bar1)
    bar1.height = 4.5
    bar1.width = 9.25
    ws.add_chart(bar1, "D15")

    bar2 = BarChart()
    bar2.type = "bar"
    bar2.style = 11
    bar2.legend = None
    data = Reference(ws, min_col=18, min_row=industry_header_row, max_row=industry_last_row)
    cats = Reference(ws, min_col=17, min_row=industry_header_row + 1, max_row=industry_last_row)
    bar2.add_data(data, titles_from_data=True)
    series = bar2.series[0]
    series.graphicalProperties.solidFill = "5B9A5B"   # same green as Family-Owned KPI
    series.graphicalProperties.line.solidFill = "5B9A5B"
    bar2.set_categories(cats)
    bar2.dataLabels = DataLabelList()
    bar2.dataLabels.showVal = True
    bar2.y_axis.majorGridlines = None
    bar2.x_axis.majorGridlines = None
    bar2.y_axis.delete = True
    bar2.x_axis.delete = True
    bar2.dataLabels.showCatName = False
    bar2.dataLabels.showSerName = False
    bar2.dataLabels.showLegendKey = False
    style_chart(bar2)
    bar2.height = 4.5
    bar2.width = 9.25
    ws.add_chart(bar2, "G15")

    pct_founder_led = round(100 * sum(1 for c in candidates if c.get("founder_led") == "Yes") / total_targets, 0) if total_targets else 0
    pct_family_owned = round(100 * sum(1 for c in candidates if c.get("family_owned") == "Yes") / total_targets, 0) if total_targets else 0
    top_state_label = top_states[0][0] if top_states else "N/A"
    high_scorers = sum(1 for c in candidates if (c.get("seller_readiness_score") or 0) >= 80)
    top_industry = top_industries[0][0] if top_industries else "N/A"

    insights_col = 10          # J
    start_row = 15

    insights = [
        ("📍", "Top Geography", top_state_label),
        ("🏭", "Top Industry", top_industry),
        ("👥", "Ownership",
        f"{pct_founder_led:.0f}% Founder / {pct_family_owned:.0f}% Family"),
        ("⭐", "High Opportunity",
        f"{high_scorers} Company" if high_scorers == 1 else f"{high_scorers} Companies"),
        ("🆕", "New This Run",
        f"{new_this_week} Company" if new_this_week == 1 else f"{new_this_week} Companies"),
    ]

    row = start_row

    for icon, heading, value in insights:

        ws.merge_cells(
            start_row=row,
            start_column=10,
            end_row=row,
            end_column=11
        )

        cell = ws.cell(row=row, column=10)

        cell.value = f"{icon}  {heading}: {value}"

        cell.font = Font(
            name=FONT_NAME,
            size=10,
            bold=False,
            color="333333"
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            indent=0
        )

        cell.fill = PatternFill(
            "solid",
            start_color="FBFBF7",
            end_color="FBFBF7"
        )

        cell.border = Border(
            bottom=Side(style="thin", color="ECECEC")
        )

        row += 1
    
    # =====================================================
    # TOP OPPORTUNITY COMPANIES SECTION
    # =====================================================
    table_title_row = 21
    ws.merge_cells(
        start_row=table_title_row,
        start_column=1,
        end_row=table_title_row,
        end_column=11
    )

    title = ws.cell(row=table_title_row, column=1)
    title.value = "TOP OPPORTUNITY COMPANIES"
    title.fill = NAVY_FILL
    title.font = Font(
        name=FONT_NAME,
        size=12,
        bold=True,
        color="FFFFFF"
    )

    title.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=1
    )

    ws.row_dimensions[table_title_row].height = 24
    table_columns = [
        ("Rank", None, 1),
        ("Company Name", "company_name", 26),
        ("City", "city", 18),
        ("Industry", "industry", 20),
        ("Why Discovered", "why_discovered", 44),
        ("Founder Name", "founder_name", 20),
        ("Ownership Status", "ownership_status", 16),
        ("Years in Business", "years_in_business", 14),
        ("Status", "review_status", 12),
        ("Fit Analysis", "fit_analysis", 44),
        ("Readiness Score", "seller_readiness_score", 13),
    ]
    header_row = 22
    for i, (label, _, _) in enumerate(table_columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 24

    sorted_candidates = sorted(
        candidates,
        key=lambda r: (r.get("first_seen_date") is None, r.get("first_seen_date")),
        reverse=True,
    )
 
    # Columns whose content is long enough to need wrapping instead of
    # spilling visually into the next cell (which is what was happening
    # before - text from Why Discovered/Fit Analysis/Website/etc bled
    # into neighboring columns because only 3 columns had wrap_text set).
    WRAP_COLUMNS = {
        "company_name", "city", "industry", "why_discovered",
        "founder_name", "ownership_status", "fit_analysis",
    }
 
    for idx, row in enumerate(sorted_candidates, start=1):
        r = header_row + idx
        rank_cell = ws.cell(row=r, column=1, value=idx)
        rank_cell.font = BOLD_FONT
        rank_cell.alignment = Alignment(horizontal="center", vertical="center")
        for c, (_, key, _) in enumerate(table_columns[1:], start=2):
            value = row.get(key)
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if key in ("years_in_business", "seller_readiness_score"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    wrap_text=(key in WRAP_COLUMNS),
                    vertical="center"
                )
            if key == "seller_readiness_score":
                fill = score_fill(value)
                if fill:
                    cell.fill = fill
                cell.font = BOLD_FONT
            if key == "review_status":
                cell.font = status_font(value)
        # Fixed row height so long Why Discovered/Fit Analysis text has
        # room to wrap instead of being clipped at default row height.
        ws.row_dimensions[r].height = 42

    ws.auto_filter.ref = (
        f"A{header_row}:"
        f"{get_column_letter(len(table_columns))}{header_row}"
    )

    autosize_columns(ws, [w for _, _, w in table_columns] + [12] * (TOTAL_COLS - len(table_columns)))

    # Equal dashboard column widths (A:I)
    DASHBOARD_COL_WIDTH = 16

    for col in range(1, 11):  # Columns A:J
        ws.column_dimensions[get_column_letter(col)].width = DASHBOARD_COL_WIDTH

    return ws
 
 
DB_COLUMNS = [
    ("Company Name", "company_name", 28), ("Website", "website", 24),
    ("City", "city", 16), ("State", "state", 12),
    ("Industry", "industry", 18), ("Company Type", "company_type", 14),
    ("Company Description", "company_description", 40),
    ("Why Discovered", "why_discovered", 34),
    ("Founded Year", "founded_year", 12), ("Years in Business", "years_in_business", 14),
    ("Founder Name", "founder_name", 20), ("Founder Led", "founder_led", 12),
    ("Family Owned", "family_owned", 12), ("Founder Age Est.", "founder_age_estimate", 14),
    ("Ownership Status", "ownership_status", 16),
    ("Fit Analysis", "fit_analysis", 40),
    ("Seller Readiness Signals", "seller_readiness_signals", 34),
    ("Seller Readiness Score", "seller_readiness_score", 16), ("Review Status", "review_status", 14),
    ("First Seen", "first_seen_date", 14), ("Last Seen", "last_seen_date", 14),
]
 
 
def build_companies_db_sheet(wb, candidates):
    ws = wb.create_sheet("Companies_DB")
    header_row = 1
    num_cols = len(DB_COLUMNS)
    for i, (label, _, _) in enumerate(DB_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 32
    DB_WRAP_COLUMNS = {
        "company_name", "city", "industry", "company_description",
        "why_discovered", "website", "founder_name", "ownership_status",
        "fit_analysis", "seller_readiness_signals",
    }

    for r, row in enumerate(candidates, start=header_row + 1):
        for c, (_, key, _) in enumerate(DB_COLUMNS, start=1):
            value = row.get(key)
            cell = ws.cell(
                row=r,
                column=c,
                value=value
            )

            if key == "website" and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                wrap_text=(key in DB_WRAP_COLUMNS),
                horizontal="left",
                vertical="center"
            )
            if key == "seller_readiness_score":
                fill = score_fill(row.get(key))
                if fill:
                    cell.fill = fill
        ws.row_dimensions[r].height = 60
        if r % 2 == 0:
            alt_fill = PatternFill(
                "solid",
                start_color="FAFAFA",
                end_color="FAFAFA"
            )

            for c in range(1, num_cols + 1):
                if DB_COLUMNS[c-1][1] != "seller_readiness_score":
                    ws.cell(r, c).fill = alt_fill
    autosize_columns(ws, [w for _, _, w in DB_COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{header_row}"
    return ws
 
 
TOP_N = 10
 
from openpyxl.worksheet.datavalidation import DataValidation
FEEDBACK_OPTIONS = ["New", "Pursuing", "Passed", "Bad Data"]
 
 
def build_top_companies_sheet(wb, candidates):
    ws = wb.create_sheet("Top_Companies")
   
    header_row = 1
    scored = [c for c in candidates if c.get("seller_readiness_score") is not None]
    top = sorted(scored, key=lambda r: -r["seller_readiness_score"])[:TOP_N]
 
    columns = [
        ("Rank", None, 6),
        ("Company Name", "company_name", 26),
        ("City", "city", 16),
        ("State", "state", 10),
        ("Industry", "industry", 16),
        ("Why Discovered", "why_discovered", 40),
        ("Founder Name", "founder_name", 20),
        ("Founder Led", "founder_led", 12),
        ("Family Owned", "family_owned", 12),
        ("Readiness Score", "seller_readiness_score", 12),
        ("Feedback", "review_status", 14),
        ("Notes", "review_notes", 35),
        ("_candidate_id", "candidate_id", 1),
    ]
 
    for i, (label, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 32
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(FEEDBACK_OPTIONS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
 
    feedback_col_idx = next(i for i, (label, _, _) in enumerate(columns, start=1) if label == "Feedback")
    feedback_col_letter = get_column_letter(feedback_col_idx)
 
    for idx, row in enumerate(top, start=1):
        r = header_row + idx 
        ws.cell(row=r, column=1, value=idx).font = BOLD_FONT
        for c, (label, key, _) in enumerate(columns[1:], start=2):
            if key == "review_status":
                value = row.get("review_status") or "New"
            elif key == "review_notes":
                value = row.get("review_notes") or ""
            else:
                value = row.get(key)
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
            if key == "seller_readiness_score":
                fill = score_fill(row.get(key))
                if fill:
                    cell.fill = fill
                cell.font = BOLD_FONT
        ws.row_dimensions[r].height = 42
        dv.add(f"{feedback_col_letter}{r}")
 
    autosize_columns(ws, [w for _, _, w in columns])
    ws.auto_filter.ref = (
        f"A{header_row}:"
        f"{get_column_letter(len(columns))}{header_row}"
    )

    id_col_idx = next(i for i, (label, _, _) in enumerate(columns, start=1) if label == "_candidate_id")
    ws.column_dimensions[get_column_letter(id_col_idx)].width = 0
    ws.column_dimensions[get_column_letter(id_col_idx)].hidden = True
 
    ws.freeze_panes = "A2"
    return ws
 
 
def generate_dashboard(output_file, search_request_id):
    """
    search_request_id scopes everything: the client only sees leads for
    the search they're currently running.
    """
    candidates = fetch_all_candidates(search_request_id)
    last_week = fetch_last_week_snapshot(search_request_id)
 
    wb = Workbook()
    wb.remove(wb.active)

    search_request = fetch_search_request(search_request_id)
    build_dashboard_sheet(
        wb,
        candidates,
        last_week,
        search_request
    )
    build_companies_db_sheet(wb, candidates)
    build_top_companies_sheet(wb, candidates)
 
    wb.save(output_file)
    print(f"Saved dashboard workbook ({len(candidates)} candidates "
          f"for search {search_request_id}) -> {output_file}")
    return output_file